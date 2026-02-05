import io
import os
import json
import pandas as pd
import streamlit as st
from datetime import time as dtime
from streamlit_calendar import calendar
import holidays

# ============================================================
# CONFIG FIJA
# ============================================================
FECHA_INICIO = "2026-03-16"   # NO cambiar
NUMERO_SEMANAS = 20          # NO cambiar
TIMEZONE = "America/Santiago"

LOGO_PATH = "assets/logo.png"   # pon tu logo aquí (crea carpeta assets/)

DATA_DIR = "data"
EXCEL_FILENAME = "calendario.xlsx"
EXCEL_PATH = os.path.join(DATA_DIR, EXCEL_FILENAME)
MISIONES_STATE_PATH = os.path.join(DATA_DIR, "misiones_estado.json")

# ============================================================
# Utilidades
# ============================================================
ORDEN_DIAS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DIAS_ES = {
    "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
    "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"
}
IDX_A_DIA_ES = {0:"Lunes",1:"Martes",2:"Miércoles",3:"Jueves",4:"Viernes",5:"Sábado",6:"Domingo"}

PROF_PALETTE = {
    # colores “amables” (puedes ajustar)
    "TY": "#1f77b4",
    "IG": "#2ca02c",
    "CC": "#ff7f0e",
    "AR": "#9467bd",
    "JCS": "#8c564b",
}

EVAL_ICON = {
    "Prueba": "⭐",
    "Experimento": "🧪",
    "Trabajo práctico": "📝",
}
MISION_ICON = {
    "Propuesta de preguntas": "💡",
    "Revisión taller": "🧰",
    "Revisión prueba": "🔎",
}

ACT_COLORS = {
    "Clase teórica": "#1f77b4",
    "Seminario": "#2ca02c",
    "Laboratorio": "#ff7f0e",
    "Trabajo autónomo": "#9467bd",
    "Sin clases (Feriado)": "#d62728",
    "Examen": "#111111",
    "Misión": "#8c564b",
}

def ensure_dirs():
    os.makedirs(DATA_DIR, exist_ok=True)

def parse_time_hhmm(s: str) -> dtime:
    h, m = s.split(":")
    return dtime(int(h), int(m))

def horario_str(inicio: dtime, fin: dtime) -> str:
    return f"{inicio.strftime('%H:%M')}–{fin.strftime('%H:%M')}"

def extraer_inicio(h: str) -> str:
    if isinstance(h, str) and "–" in h:
        return h.split("–")[0]
    return "00:00"

def split_profes(s: str):
    if not s:
        return []
    return [x.strip() for x in str(s).split(",") if x.strip()]

def row_has_prof(row_prof: str, selected_set: set) -> bool:
    # Si el row no tiene profesores, NO pasa el filtro cuando se filtra por profe
    profs = set(split_profes(row_prof))
    if not profs:
        return False
    return len(profs.intersection(selected_set)) > 0

def fecha_de_dia_en_semana(lunes_semana: pd.Timestamp, dia_semana: str) -> pd.Timestamp:
    idx = ORDEN_DIAS.index(dia_semana)  # Monday=0 ... Sunday=6
    return pd.Timestamp(lunes_semana.date()) + pd.Timedelta(days=idx)

def feriados_chile_entre(fecha_min, fecha_max):
    fmin = pd.Timestamp(fecha_min).date()
    fmax = pd.Timestamp(fecha_max).date()
    years = list(range(fmin.year, fmax.year + 1))
    cl = holidays.country_holidays("CL", years=years)
    lista = []
    for d, nombre in cl.items():
        if fmin <= d <= fmax:
            lista.append({"fecha": pd.Timestamp(d).strftime("%Y-%m-%d"), "nombre": str(nombre)})
    return sorted(lista, key=lambda x: x["fecha"])

# ============================================================
# Generador de calendario (CLASES) - NO cambies horarios/fechas
# ============================================================
def crear_calendario_curso(
    fecha_inicio=FECHA_INICIO,
    numero_semanas=NUMERO_SEMANAS,
):
    # Horarios definitivos (los que dijiste que ya están bien)
    secciones = {
        "Sección 1": {
            "teorica":   {"dia_semana": "Monday",    "inicio": dtime(12, 0),  "fin": dtime(13, 30)},
            "seminario": {"dia_semana": "Wednesday", "inicio": dtime(15, 0),  "fin": dtime(16, 30)},
            "lab":       {"dia_semana": "Wednesday", "inicio": dtime(16, 45), "fin": dtime(18, 15)},
        },
        "Sección 2": {
            "teorica":   {"dia_semana": "Monday", "inicio": dtime(16, 45), "fin": dtime(18, 15)},
            "seminario": {"dia_semana": "Friday", "inicio": dtime(8, 30),  "fin": dtime(10, 0)},
            "lab":       {"dia_semana": "Monday", "inicio": dtime(15, 0),  "fin": dtime(16, 30)},
        },
        "Sección 3": {
            "teorica":   {"dia_semana": "Wednesday", "inicio": dtime(16, 45), "fin": dtime(18, 15)},
            "seminario": {"dia_semana": "Friday",    "inicio": dtime(10, 15), "fin": dtime(11, 45)},
            "lab":       {"dia_semana": "Friday",    "inicio": dtime(12, 0),  "fin": dtime(13, 30)},
        },
        "Sección 4": {
            "teorica":   {"dia_semana": "Wednesday", "inicio": dtime(15, 0),  "fin": dtime(16, 30)},
            "seminario": {"dia_semana": "Friday",    "inicio": dtime(8, 30),  "fin": dtime(10, 0)},
            "lab":       {"dia_semana": "Friday",    "inicio": dtime(10, 15), "fin": dtime(11, 45)},
        },
    }

    fecha_inicio = pd.Timestamp(fecha_inicio)
    filas = []

    for seccion, cfg in secciones.items():
        for semana in range(1, numero_semanas + 1):
            lunes = fecha_inicio + pd.Timedelta(days=7*(semana-1))

            # Teórica + Seminario
            for clave, nombre in [("teorica","Clase teórica"), ("seminario","Seminario")]:
                fecha_evento = fecha_de_dia_en_semana(lunes, cfg[clave]["dia_semana"]).date()
                horario = horario_str(cfg[clave]["inicio"], cfg[clave]["fin"])
                filas.append({
                    "semana": semana,
                    "fecha": fecha_evento,
                    "día": DIAS_ES[cfg[clave]["dia_semana"]],
                    "horario": horario,
                    "sección": seccion,
                    "actividad": nombre,
                    "tema": "",
                    "evaluación": "",
                    "profesores": "",
                    "misiones": "",
                    "observaciones": ""
                })

            # Laboratorio semanas pares
            if semana % 2 == 0:
                fecha_evento = fecha_de_dia_en_semana(lunes, cfg["lab"]["dia_semana"]).date()
                horario = horario_str(cfg["lab"]["inicio"], cfg["lab"]["fin"])
                filas.append({
                    "semana": semana,
                    "fecha": fecha_evento,
                    "día": DIAS_ES[cfg["lab"]["dia_semana"]],
                    "horario": horario,
                    "sección": seccion,
                    "actividad": "Laboratorio",
                    "tema": "",
                    "evaluación": "",
                    "profesores": "",
                    "misiones": "",
                    "observaciones": "Cada 2 semanas"
                })

    df = pd.DataFrame(filas)
    df["_inicio_dt"] = pd.to_datetime(df["fecha"].astype(str) + " " + df["horario"].apply(extraer_inicio))
    df = df.sort_values(["sección", "semana", "_inicio_dt"]).drop(columns=["_inicio_dt"])
    return df

# ============================================================
# CASOS ESPECIALES (EDITA AQUÍ)
# ============================================================
CASOS_ESPECIALES = {
    "ediciones": [
        # demo para que no se vea vacío
        {"filtro": {"sección":"Sección 1", "semana":1, "actividad":"Clase teórica"},
         "set": {"tema":"Trigonometría", "profesores":"TY", "observaciones":"Introducción"}},
        {"filtro": {"sección":"Sección 1", "semana":1, "actividad":"Seminario"},
         "set": {"tema":"Lógica 1", "profesores":"IG"}},
        {"filtro": {"sección":"Sección 4", "semana":8, "actividad":"Clase teórica"},
         "set": {"evaluación":"Prueba", "profesores":"IG", "observaciones":"Duración: 45 min"}},
    ],

    "semanas_trabajo_autonomo": [16],

    "semanas_examenes": [17, 18],
    "eventos_examenes": [
        {"sección": "Sección 1", "fecha": "2026-07-06", "inicio": dtime(9,0),  "fin": dtime(11,0),
         "actividad": "Examen", "profesores": "TY", "observaciones": "Examen final"},
        {"sección": "Sección 2", "fecha": "2026-07-07", "inicio": dtime(16,0), "fin": dtime(18,0),
         "actividad": "Examen", "profesores": "IG", "observaciones": "Examen final"},
        {"sección": "Sección 3", "fecha": "2026-07-08", "inicio": dtime(10,0), "fin": dtime(12,0),
         "actividad": "Examen", "profesores": "CC", "observaciones": "Examen final"},
        {"sección": "Sección 4", "fecha": "2026-07-09", "inicio": dtime(18,0), "fin": dtime(20,0),
         "actividad": "Examen", "profesores": "AR, JCS", "observaciones": "Examen final"},
    ],

    # Misiones (all-day, no bloquean el horario)
    "eventos_misiones": [
        {"fecha":"2026-04-10", "sección":"Equipo docente", "misiones":"Propuesta de preguntas",
         "profesores":"TY, IG", "observaciones":"Subir documento a carpeta compartida"},
        {"fecha":"2026-05-08", "sección":"Equipo docente", "misiones":"Revisión prueba",
         "profesores":"IG", "observaciones":"Enviar versión final"},
        {"fecha":"2026-06-05", "sección":"Equipo docente", "misiones":"Revisión taller",
         "profesores":"CC, AR", "observaciones":"Revisar guía y pauta"},
    ],
}

def aplicar_casos_especiales(df: pd.DataFrame, reglas: dict) -> pd.DataFrame:
    df = df.copy()

    # A) Ediciones puntuales
    for ed in reglas.get("ediciones", []):
        filtro = ed.get("filtro", {})
        mask = pd.Series(True, index=df.index)
        for k, v in filtro.items():
            mask &= (df[k] == v)
        for col, val in ed.get("set", {}).items():
            df.loc[mask, col] = val

    # B) Trabajo autónomo
    semanas_auto = set(reglas.get("semanas_trabajo_autonomo", []))
    if semanas_auto:
        mask_auto = df["semana"].isin(semanas_auto)
        df.loc[mask_auto, "actividad"] = "Trabajo autónomo"
        df.loc[mask_auto, "observaciones"] = "No hay clases (trabajo autónomo)."

    # C) Feriados automáticos Chile
    fer = feriados_chile_entre(df["fecha"].min(), df["fecha"].max())
    fer_map = {pd.Timestamp(f["fecha"]).date(): f["nombre"] for f in fer}
    mask_fer = df["fecha"].isin(list(fer_map.keys()))
    df.loc[mask_fer, "actividad"] = "Sin clases (Feriado)"
    df.loc[mask_fer, "observaciones"] = df.loc[mask_fer, "fecha"].map(fer_map)

    # D) Exámenes: borra clases regulares en esas semanas + inserta exámenes
    semanas_ex = set(reglas.get("semanas_examenes", []))
    if semanas_ex:
        ref = pd.Timestamp(str(df["fecha"].min()))
        df = df[~df["semana"].isin(semanas_ex)].copy()

        nuevos = []
        for ev in reglas.get("eventos_examenes", []):
            fecha_ts = pd.Timestamp(ev["fecha"])
            semana_ev = int((fecha_ts - ref).days // 7) + 1
            nuevos.append({
                "semana": semana_ev,
                "fecha": fecha_ts.date(),
                "día": IDX_A_DIA_ES[fecha_ts.weekday()],
                "horario": horario_str(ev["inicio"], ev["fin"]),
                "sección": ev["sección"],
                "actividad": ev["actividad"],
                "tema": ev.get("tema", ""),
                "evaluación": ev.get("evaluación", ""),
                "profesores": ev.get("profesores", ""),
                "misiones": "",
                "observaciones": ev.get("observaciones", ""),
            })
        if nuevos:
            df = pd.concat([df, pd.DataFrame(nuevos)], ignore_index=True)

    # E) Misiones: inserta filas all-day (horario vacío)
    misiones = reglas.get("eventos_misiones", [])
    if misiones:
        ref = pd.Timestamp(str(df["fecha"].min()))
        nuevos = []
        for m in misiones:
            fecha_ts = pd.Timestamp(m["fecha"])
            semana_ev = int((fecha_ts - ref).days // 7) + 1
            nuevos.append({
                "semana": semana_ev,
                "fecha": fecha_ts.date(),
                "día": IDX_A_DIA_ES[fecha_ts.weekday()],
                "horario": "",  # all-day
                "sección": m.get("sección", "Equipo docente"),
                "actividad": "Misión",
                "tema": "",
                "evaluación": "",
                "profesores": m.get("profesores", ""),
                "misiones": m.get("misiones", ""),
                "observaciones": m.get("observaciones", ""),
            })
        df = pd.concat([df, pd.DataFrame(nuevos)], ignore_index=True)

    # Orden
    df["_inicio_dt"] = pd.to_datetime(
        df["fecha"].astype(str) + " " + df["horario"].fillna("").apply(extraer_inicio),
        errors="coerce"
    )
    df = df.sort_values(["fecha", "_inicio_dt", "sección"], na_position="last").drop(columns=["_inicio_dt"])
    return df

# ============================================================
# Excel export robusto (evita 70348 y ######)
# ============================================================
def df_a_excel_bytes(df: pd.DataFrame) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl", datetime_format="DD/MM/YYYY", date_format="DD/MM/YYYY") as writer:
        df2 = df.copy()
        # Aseguramos datetime/date
        df2["fecha"] = pd.to_datetime(df2["fecha"]).dt.date
        df2.to_excel(writer, index=False, sheet_name="Calendario")

        ws = writer.sheets["Calendario"]

        # Ajuste simple de columnas (ancho + formato)
        header = [c.value for c in ws[1]]
        col_index = {name: i+1 for i, name in enumerate(header)}

        # formato fecha
        if "fecha" in col_index:
            c = col_index["fecha"]
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c).number_format = "DD/MM/YYYY"

        # wrap + align
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # auto-width razonable
        for i, name in enumerate(header, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(28, max(12, len(str(name)) + 2))

    return buffer.getvalue()

# ============================================================
# FullCalendar events
# ============================================================
def df_a_fullcalendar_events(df: pd.DataFrame):
    events = []

    for _, r in df.iterrows():
        fecha = pd.Timestamp(str(r["fecha"]))
        horario = str(r.get("horario", "") or "").strip()

        tema = str(r.get("tema", "") or "").strip()
        evaluacion = str(r.get("evaluación", "") or "").strip()
        profs = str(r.get("profesores", "") or "").strip()
        obs = str(r.get("observaciones", "") or "").strip()
        mision_txt = str(r.get("misiones", "") or "").strip()
        actividad = str(r.get("actividad", "") or "").strip()

        # título: prioriza mostrar tema; y si hay evaluación, icono
        eval_ic = EVAL_ICON.get(evaluacion, "")
        prefix = (eval_ic + " ") if eval_ic else ""

        if actividad == "Misión":
            ic = MISION_ICON.get(mision_txt, "📌")
            title = f"{ic} {mision_txt}".strip()
            start = fecha.date().isoformat()
            end = (fecha + pd.Timedelta(days=1)).date().isoformat()
            events.append({
                "title": title,
                "start": start,
                "end": end,
                "allDay": True,
                "color": ACT_COLORS.get("Misión", "#888888"),
                "extendedProps": {
                    "tipo": "mision",
                    "semana": r.get("semana", ""),
                    "sección": r.get("sección", ""),
                    "misiones": mision_txt,
                    "profesores": profs,
                    "observaciones": obs,
                }
            })
            continue

        # Eventos con horario
        all_day = False
        if "–" in horario:
            a, b = horario.split("–")
            hi = pd.to_datetime(a, format="%H:%M").time()
            hf = pd.to_datetime(b, format="%H:%M").time()
            start_dt = fecha + pd.Timedelta(hours=hi.hour, minutes=hi.minute)
            end_dt = fecha + pd.Timedelta(hours=hf.hour, minutes=hf.minute)
            start = start_dt.isoformat()
            end = end_dt.isoformat()
        else:
            # fallback
            all_day = True
            start = fecha.date().isoformat()
            end = (fecha + pd.Timedelta(days=1)).date().isoformat()

        # Mostrar tema en el título si existe
        if tema:
            title = f"{prefix}{actividad} · {tema}"
        else:
            title = f"{prefix}{actividad}"

        # si hay evaluación, destacar con color distinto (pero simple: usamos color actividad)
        color = ACT_COLORS.get(actividad, "#888888")

        events.append({
            "title": title,
            "start": start,
            "end": end,
            "allDay": all_day,
            "color": color,
            "extendedProps": {
                "tipo": "clase",
                "semana": r.get("semana", ""),
                "día": r.get("día", ""),
                "horario": horario,
                "sección": r.get("sección", ""),
                "actividad": actividad,
                "tema": tema,
                "evaluación": evaluacion,
                "profesores": profs,
                "observaciones": obs,
            }
        })

    return events

# ============================================================
# Misiones: estado (completada) persistente simple
# ============================================================
def load_misiones_state() -> dict:
    ensure_dirs()
    if not os.path.exists(MISIONES_STATE_PATH):
        return {}
    try:
        with open(MISIONES_STATE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_misiones_state(state: dict):
    ensure_dirs()
    with open(MISIONES_STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def mission_uid(row: pd.Series) -> str:
    # uid estable para guardar el checkbox
    return f"{row.get('fecha')}|{row.get('misiones')}|{row.get('sección')}|{row.get('profesores')}"

def tag_persona(code: str) -> str:
    color = PROF_PALETTE.get(code, "#777777")
    return f"""
    <span style="
        display:inline-block;
        padding:2px 8px;
        border-radius:999px;
        background:{color};
        color:white;
        font-size:12px;
        margin-right:6px;">
        {code}
    </span>
    """

# ============================================================
# UI
# ============================================================
st.set_page_config(page_title="Calendario del Curso", layout="wide")

# Header con logo (derecha)
c1, c2 = st.columns([5, 1])
with c1:
    st.title("📅 Calendario del Curso")
    st.caption("Calendario + tabla Excel + misiones + métricas para el equipo docente.")
with c2:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, use_container_width=True)

# Construir DF base
ensure_dirs()
df = crear_calendario_curso(FECHA_INICIO, NUMERO_SEMANAS)
df = aplicar_casos_especiales(df, CASOS_ESPECIALES)

# Opcional: guardar un excel “oficial” en disco (sirve para inspección local)
# (En Streamlit Cloud esto puede no ser persistente, pero no molesta)
try:
    bytes_xlsx_full = df_a_excel_bytes(df)
    with open(EXCEL_PATH, "wb") as f:
        f.write(bytes_xlsx_full)
except Exception:
    pass

# Valores para filtros (checkbox)
all_secciones = sorted(df["sección"].unique())
all_prof_codes = sorted({p for s in df["profesores"].dropna().unique() for p in split_profes(s)})

# Estado inicial en session_state
if "sel_secciones" not in st.session_state:
    st.session_state.sel_secciones = {s: True for s in all_secciones}
if "sel_profes" not in st.session_state:
    st.session_state.sel_profes = {p: True for p in all_prof_codes}

# Tabs
tab1, tab2, tab3, tab4 = st.tabs(["📆 Calendario", "📄 Excel", "🧭 Misiones", "📊 Plots"])

# ============================================================
# TAB 1: Calendario
# ============================================================
with tab1:
    st.subheader("📆 Calendario")

    # Índice horizontal para “ir a semana”
    weeks = list(range(1, NUMERO_SEMANAS + 1))
    semana_sel = st.radio(
        "Ir a semana (del curso)",
        options=weeks,
        horizontal=True,
        index=0,
        key="nav_semana"
    )

    st.markdown("---")

    # Filtros por checkboxes (no sidebar)
    fcol1, fcol2 = st.columns([1, 1])

    with fcol1:
        st.markdown("### Secciones")
        for s in all_secciones:
            st.session_state.sel_secciones[s] = st.checkbox(
                s, value=st.session_state.sel_secciones.get(s, True), key=f"sec_{s}"
            )

        if st.button("✅ Marcar todas (Secciones)"):
            for s in all_secciones:
                st.session_state.sel_secciones[s] = True
                st.session_state[f"sec_{s}"] = True

        if st.button("❌ Desmarcar todas (Secciones)"):
            for s in all_secciones:
                st.session_state.sel_secciones[s] = False
                st.session_state[f"sec_{s}"] = False

    with fcol2:
        st.markdown("### Profesores")
        if not all_prof_codes:
            st.info("Aún no hay profesores cargados en el calendario.")
        else:
            for p in all_prof_codes:
                st.session_state.sel_profes[p] = st.checkbox(
                    p, value=st.session_state.sel_profes.get(p, True), key=f"prof_{p}"
                )

            if st.button("✅ Marcar todos (Profesores)"):
                for p in all_prof_codes:
                    st.session_state.sel_profes[p] = True
                    st.session_state[f"prof_{p}"] = True

            if st.button("❌ Desmarcar todos (Profesores)"):
                for p in all_prof_codes:
                    st.session_state.sel_profes[p] = False
                    st.session_state[f"prof_{p}"] = False

    # Construir filtros seleccionados
    sec_selected = {s for s, ok in st.session_state.sel_secciones.items() if ok}
    prof_selected = {p for p, ok in st.session_state.sel_profes.items() if ok}

    df_f = df[df["sección"].isin(sec_selected)].copy()
    if all_prof_codes:
        # mantener solo filas que tengan algún prof seleccionado
        df_f = df_f[df_f["profesores"].apply(lambda x: row_has_prof(x, prof_selected) if prof_selected else True)].copy()

    # Initial date: lunes de la semana seleccionada
    initial_date = (pd.Timestamp(FECHA_INICIO) + pd.Timedelta(days=7*(semana_sel-1))).strftime("%Y-%m-%d")

    # FullCalendar
    events = df_a_fullcalendar_events(df_f)

    calendar_options = {
        "initialView": "timeGridWeek",
        "initialDate": initial_date,
        "headerToolbar": {
            "left": "prev,next today",
            "center": "title",
            "right": "dayGridMonth,timeGridWeek,timeGridDay,listWeek",
        },
        "height": 850,
        "contentHeight": "auto",
        "slotMinTime": "07:00:00",
        "slotMaxTime": "22:00:00",
        "scrollTime": "08:00:00",
        "slotDuration": "00:30:00",
        "snapDuration": "00:15:00",
        "selectable": True,
        "editable": False,
        "nowIndicator": True,
        "weekNumbers": True,
        "dayMaxEvents": True,
    }

    state = calendar(events=events, options=calendar_options)

    st.divider()
    st.subheader("🔎 Detalles (clic en un evento)")
    event_click = state.get("eventClick") if isinstance(state, dict) else None
    if event_click:
        ev = event_click.get("event", {})
        props = ev.get("extendedProps", {})
        st.write({
            "Título": ev.get("title", ""),
            "Inicio": ev.get("start", ""),
            "Fin": ev.get("end", ""),
            "Tipo": props.get("tipo", ""),
            "Sección": props.get("sección", ""),
            "Actividad": props.get("actividad", ""),
            "Misión": props.get("misiones", ""),
            "Tema": props.get("tema", ""),
            "Evaluación": props.get("evaluación", ""),
            "Profesores": props.get("profesores", ""),
            "Observaciones": props.get("observaciones", ""),
            "Semana": props.get("semana", ""),
        })
    else:
        st.caption("Clickea una clase o misión para ver detalles.")

# ============================================================
# TAB 2: Excel (tabla + descargar xlsx)
# ============================================================
with tab2:
    st.subheader("📄 Vista tipo Excel")

    # Usamos df completo o df filtrado? (yo dejaría completo y que el filtro sea visual)
    show_df = df.copy()
    show_df["fecha"] = pd.to_datetime(show_df["fecha"])  # para buen render

    st.dataframe(
        show_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "fecha": st.column_config.DateColumn(format="DD/MM/YYYY"),
        }
    )

    st.markdown("### ⬇️ Descargar Excel (.xlsx)")
    excel_bytes = df_a_excel_bytes(df)
    st.download_button(
        label="Descargar calendario.xlsx",
        data=excel_bytes,
        file_name="calendario.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================================
# TAB 3: Misiones (lista + completadas)
# ============================================================
with tab3:
    st.subheader("🧭 Misiones")

    mis_df = df[df["actividad"] == "Misión"].copy()
    if mis_df.empty:
        st.info("No hay misiones configuradas aún.")
    else:
        state = load_misiones_state()

        # ordenar por fecha
        mis_df["fecha_dt"] = pd.to_datetime(mis_df["fecha"])
        mis_df = mis_df.sort_values("fecha_dt").drop(columns=["fecha_dt"])

        # construir lista con estado
        filas = []
        for _, row in mis_df.iterrows():
            uid = mission_uid(row)
            done = bool(state.get(uid, False))
            filas.append((done, uid, row))

        # separa pendientes vs completadas
        pendientes = [x for x in filas if not x[0]]
        completadas = [x for x in filas if x[0]]

        # def render_group(title, items, gray=False):
        #     st.markdown(f"### {title}")
        #     if not items:
        #         st.caption("—")
        #         return
        #     for done, uid, row in items:
        #         fecha = pd.Timestamp(row["fecha"]).strftime("%d/%m/%Y")
        #         mtxt = str(row.get("misiones", "") or "").strip()
        #         profs = split_profes(row.get("profesores", ""))

        #         # checkbox
        #         key = f"m_done_{uid}"
        #         if key not in st.session_state:
        #             st.session_state[key] = done

        #         cA, cB = st.columns([1, 8])
        #         with cA:
        #             st.session_state[key] = st.checkbox("", value=st.session_state[key], key=key)
        #         with cB:
        #             icon = MISION_ICON.get(mtxt, "📌")
        #             prof_tags = "".join(tag_persona(p) for p in profs) if profs else "<span style='color:#777'>Sin asignar</span>"
        #             style = "color:#777;" if gray else ""
        #             st.markdown(
        #                 f"""
        #                 <div style="{style}">
        #                   <b>{icon} {mtxt}</b> — <span style="opacity:0.9">{fecha}</span><br/>
        #                   {prof_tags}<br/>
        #                   <span style="opacity:0.8">{row.get('observaciones','')}</span>
        #                 </div>
        #                 """,
        #                 unsafe_allow_html=True
        #             )

        #         # actualizar estado persistente
        #         new_done = bool(st.session_state[key])
        #         if new_done != done:
        #             state[uid] = new_done
        #             save_misiones_state(state)

        # render_group("Pendientes", pendientes, gray=False)
        # st.divider()
        # render_group("Completadas", completadas, gray=True)
        
        def render_group(title, items, gray=False):
            st.markdown(f"### {title}")
            if not items:
                st.caption("—")
                return

            for done, uid, row in items:
                fecha = pd.Timestamp(row["fecha"]).strftime("%d/%m/%Y")
                mtxt = str(row.get("misiones", "") or "").strip()
                profs = split_profes(row.get("profesores", ""))

                # checkbox: usamos una key estable, pero NO asignamos a session_state manualmente
                widget_key = f"m_done_{uid}"

                cA, cB = st.columns([1, 8])
                with cA:
                    # ✅ Streamlit gestiona session_state[widget_key]
                    st.checkbox("", value=done, key=widget_key)

                with cB:
                    icon = MISION_ICON.get(mtxt, "📌")
                    prof_tags = "".join(tag_persona(p) for p in profs) if profs else "<span style='color:#777'>Sin asignar</span>"
                    style = "color:#777;" if gray else ""
                    st.markdown(
                        f"""
                        <div style="{style}">
                        <b>{icon} {mtxt}</b> — <span style="opacity:0.9">{fecha}</span><br/>
                        {prof_tags}<br/>
                        <span style="opacity:0.8">{row.get('observaciones','')}</span>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                # ✅ leer el valor luego
                new_done = bool(st.session_state.get(widget_key, done))
                if new_done != done:
                    state[uid] = new_done
                    save_misiones_state(state)

# ============================================================
# TAB 4: Plots (sin Plotly)
# ============================================================
with tab4:
    st.subheader("📊 Plots útiles para profesores")

    # 1) Eventos por semana y tipo (conteo)
    df_plot = df.copy()
    df_plot["semana"] = pd.to_numeric(df_plot["semana"], errors="coerce")
    df_plot = df_plot.dropna(subset=["semana"])

    # Conteo por semana
    c1, c2 = st.columns(2)

    with c1:
        st.markdown("### Eventos por semana")
        week_counts = df_plot.groupby("semana").size().rename("eventos")
        st.bar_chart(week_counts)

    with c2:
        st.markdown("### Eventos por actividad")
        act_counts = df_plot.groupby("actividad").size().rename("eventos").sort_values(ascending=False)
        st.bar_chart(act_counts)

    st.divider()

    # 2) Carga por profesor (cuántos eventos tiene asignado)
    st.markdown("### Carga por profesor (eventos asignados)")
    rows = []
    for _, r in df_plot.iterrows():
        for p in split_profes(r.get("profesores", "")):
            rows.append({"profesor": p, "actividad": r.get("actividad",""), "semana": r.get("semana", None)})

    if rows:
        prof_df = pd.DataFrame(rows)
        prof_counts = prof_df.groupby("profesor").size().rename("eventos").sort_values(ascending=False)
        st.bar_chart(prof_counts)
    else:
        st.info("No hay profesores asignados aún en el calendario (columna 'profesores').")

    st.divider()

    # 3) Misiones por profesor (pendientes / completadas)
    st.markdown("### Misiones por profesor (pendientes/completadas)")
    mis_df = df[df["actividad"] == "Misión"].copy()
    if mis_df.empty:
        st.caption("No hay misiones.")
    else:
        mstate = load_misiones_state()
        rows = []
        for _, r in mis_df.iterrows():
            uid = mission_uid(r)
            done = bool(mstate.get(uid, False))
            for p in split_profes(r.get("profesores","")):
                rows.append({"profesor": p, "estado": "Completada" if done else "Pendiente"})
        if rows:
            m = pd.DataFrame(rows)
            pivot = m.pivot_table(index="profesor", columns="estado", aggfunc="size", fill_value=0)
            st.dataframe(pivot, use_container_width=True)
        else:
            st.caption("No hay profesores asignados en misiones.")