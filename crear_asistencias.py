import os
import re
import shutil
import subprocess
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ============================================================
# CONFIG
# ============================================================

DATA_DIR = "data"

CURSOS = {
    "fokito": {
        "label": "Fokito",
        "carpeta": "fokito",
    },
    "tecnologia_medica": {
        "label": "Tecnología Médica",
        "carpeta": "tecnologia_medica",
    },
    "medicina": {
        "label": "Medicina",
        "carpeta": "medicina",
    },
    "enobnu": {
        "label": "Enobnu",
        "carpeta": "enobnu",
    },
}

ACTIVIDADES_OBJETIVO = ["Seminario", "Laboratorio"]

PROFES_VALIDOS = {
    "TY", "IG", "CC", "AR", "JCS", "MB", "GM", "VB", "NV",
    "JM", "EG", "RL", "DH", "SM", "RM", "XX"
}

MAPEO_PROFES_NOMBRES = {
    "TY": ["tomás yáñez", "tomas yanez", "joriv tomás", "joriv tomas", "tomas", "tomás"],
    "IG": ["ingrid galaz", "ingrid"],
    "CC": ["caroll", "carol", "cuellar"],
    "AR": ["alexander", "riquelme"],
    "JCS": ["juan", "salas"],
    "MB": ["maximiliano", "bernal", "max"],
    "GM": ["gabriela", "martinez", "gabriela martinez"],
    "VB": ["valeria", "brancacho"],
    "NV": ["naty", "nataly", "natalia", "varas"],
    "JM": ["jose", "josé", "mondaca"],
    "EG": ["eduardo", "guerra"],
    "RL": ["rosa"],
    "DH": ["diego", "hidalgo"],
    "SM": ["sebastian", "sebastián", "sandoval"],
    "RM": ["rm"],
    "XX": ["xx"],
}

# ============================================================
# HELPERS GENERALES
# ============================================================

def limpiar_texto(x):
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()

def split_profes(valor):
    txt = limpiar_texto(valor)
    if not txt:
        return []
    return [x.strip() for x in txt.split(",") if x.strip()]

def slugify(s):
    s = limpiar_texto(s)
    s = s.lower()
    s = re.sub(r"[^\w\s-]", "", s, flags=re.UNICODE)
    s = re.sub(r"\s+", "_", s)
    return s

def asegurar_directorio(path):
    if path and not os.path.exists(path):
        os.makedirs(path)

def normalizar_columna_texto(s):
    s = limpiar_texto(s).lower()
    s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    return s

def fecha_str(x):
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    try:
        return pd.to_datetime(x).strftime("%d/%m")
    except Exception:
        return limpiar_texto(x)

def sheet_name_seguro(nombre):
    nombre = limpiar_texto(nombre)
    inval = r'[:\\/?*\[\]]'
    nombre = re.sub(inval, "", nombre)
    if len(nombre) > 31:
        nombre = nombre[:31]
    return nombre if nombre else "Hoja"

def normalizar_columnas(df):
    if df.empty:
        return df

    rename_map = {}
    for col in df.columns:
        base = limpiar_texto(col)
        lower = base.lower()

        if lower == "seccion":
            rename_map[col] = "sección"
        elif lower == "evaluacion":
            rename_map[col] = "evaluación"
        elif lower == "dia":
            rename_map[col] = "día"
        else:
            rename_map[col] = base

    return df.rename(columns=rename_map)

# ============================================================
# LECTURA CALENDARIO
# ============================================================

def cargar_calendario_excel(path_excel):
    if not os.path.exists(path_excel):
        return pd.DataFrame()

    try:
        df = pd.read_excel(path_excel, sheet_name="Calendario")
    except Exception:
        return pd.DataFrame()

    df = normalizar_columnas(df)

    for col in ["fecha", "sección", "actividad", "profesores", "tema"]:
        if col not in df.columns:
            df[col] = ""

    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df["sección"] = df["sección"].apply(limpiar_texto)
    df["actividad"] = df["actividad"].apply(limpiar_texto)
    df["profesores"] = df["profesores"].apply(limpiar_texto)
    df["tema"] = df["tema"].apply(limpiar_texto)

    return df

def extraer_eventos_asistencia(df_cal):
    if df_cal.empty:
        return pd.DataFrame()

    df = df_cal.copy()
    df = df[df["actividad"].isin(ACTIVIDADES_OBJETIVO)].copy()
    df = df.dropna(subset=["fecha"]).copy()

    filas = []

    for _, row in df.iterrows():
        fecha = row["fecha"]
        seccion = limpiar_texto(row["sección"])
        actividad = limpiar_texto(row["actividad"])
        tema = limpiar_texto(row.get("tema", ""))
        profes = split_profes(row.get("profesores", ""))

        if not seccion or not actividad:
            continue

        if not profes:
            profes = ["SIN_PROFE"]

        for profesor in profes:
            filas.append({
                "fecha": fecha,
                "sección": seccion,
                "actividad": actividad,
                "tema": tema,
                "profesor": profesor
            })

    df_out = pd.DataFrame(filas)
    if df_out.empty:
        return df_out

    df_out = df_out.sort_values(["sección", "actividad", "profesor", "fecha"]).reset_index(drop=True)

    df_out["n_clase"] = (
        df_out.groupby(["sección", "actividad", "profesor"])
        .cumcount() + 1
    )

    return df_out

# ============================================================
# LECTURA ALUMNOS DESDE CSV
# ============================================================

def detectar_profesor_desde_titulo(txt):
    txt_norm = normalizar_columna_texto(txt)

    for codigo, variantes in MAPEO_PROFES_NOMBRES.items():
        for v in variantes:
            if v in txt_norm:
                return codigo

    m = re.search(r"\b(" + "|".join(PROFES_VALIDOS) + r")\b", txt.upper())
    if m:
        return m.group(1)

    return None

def extraer_grupo_desde_titulo(txt):
    m = re.search(r"grupo\s+(\d+)", limpiar_texto(txt), flags=re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None

def inferir_seccion_para_profesor(df_eventos, profesor):
    dfp = df_eventos[df_eventos["profesor"] == profesor].copy()
    if dfp.empty:
        return None

    conteo = dfp.groupby("sección").size().sort_values(ascending=False)
    if conteo.empty:
        return None

    return str(conteo.index[0])

def cargar_alumnos_desde_integrantes_csv(ruta_csv, df_eventos):
    """
    Lee CSV tipo:
    Grupo Seminario - Grupo 1 (Prof. Tomás Yáñez)
    1, Alumno...
    2, Alumno...
    """
    if not os.path.exists(ruta_csv):
        return {}

    try:
        df_raw = pd.read_csv(ruta_csv)
    except Exception:
        try:
            df_raw = pd.read_csv(ruta_csv, sep=";")
        except Exception as e:
            print(f"⚠️ No pude leer CSV {ruta_csv}: {e}")
            return {}

    if df_raw.empty:
        return {}

    cols = list(df_raw.columns)
    col_num = cols[0] if len(cols) >= 1 else None
    col_nombre = cols[1] if len(cols) >= 2 else None

    if col_num is None or col_nombre is None:
        return {}

    resultado = {}

    profesor_actual = None
    alumnos_actuales = []

    def cerrar_bloque():
        nonlocal profesor_actual, alumnos_actuales
        if profesor_actual and alumnos_actuales:
            seccion = inferir_seccion_para_profesor(df_eventos, profesor_actual)
            if seccion:
                resultado[(seccion, profesor_actual)] = alumnos_actuales[:]
        profesor_actual = None
        alumnos_actuales = []

    for _, row in df_raw.iterrows():
        num_txt = limpiar_texto(row.get(col_num, ""))
        nom_txt = limpiar_texto(row.get(col_nombre, ""))

        if not num_txt and not nom_txt:
            cerrar_bloque()
            continue

        # encabezado tipo "Grupo Seminario - Grupo 1 (Prof. Tomás Yáñez)"
        if "grupo" in normalizar_columna_texto(num_txt) or "prof." in normalizar_columna_texto(num_txt):
            cerrar_bloque()
            profesor_detectado = detectar_profesor_desde_titulo(num_txt)
            if profesor_detectado:
                profesor_actual = profesor_detectado
            continue

        # fila de alumno: primera columna numérica
        if profesor_actual:
            if re.match(r"^\d+$", num_txt):
                nombre = limpiar_texto(nom_txt)
                if nombre:
                    alumnos_actuales.append(nombre)

    cerrar_bloque()
    return resultado

# ============================================================
# LECTURA ALUMNOS DESDE archivos alumnos_1.ods/xls/xlsx
# ============================================================

def extraer_seccion_desde_nombre_archivo(ruta_archivo):
    nombre = os.path.basename(ruta_archivo)
    m = re.search(r"alumnos_(\d+)", nombre, flags=re.IGNORECASE)
    if m:
        return f"Sección {int(m.group(1))}"
    return None

def es_fila_vacia(row):
    for val in row:
        if limpiar_texto(val) != "":
            return False
    return True

def normalizar_nombre_alumno(txt):
    txt = limpiar_texto(txt)
    txt = re.sub(r"^\s*\d+\s*[\.\)\-]?\s*", "", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt

def detectar_profesor_en_fila(txt):
    txt = limpiar_texto(txt).upper()

    if txt in PROFES_VALIDOS:
        return txt

    m = re.search(r"\b(" + "|".join(PROFES_VALIDOS) + r")\b", txt)
    if m:
        return m.group(1)

    return None

def leer_alumnos_desde_archivo_por_bloques(ruta_archivo):
    ext = os.path.splitext(ruta_archivo)[1].lower()

    if ext == ".ods":
        df_raw = pd.read_excel(ruta_archivo, sheet_name=0, header=None, engine="odf")
    else:
        df_raw = pd.read_excel(ruta_archivo, sheet_name=0, header=None)

    seccion_forzada = extraer_seccion_desde_nombre_archivo(ruta_archivo)
    if seccion_forzada is None:
        raise ValueError(
            "No pude inferir la sección desde el nombre del archivo. "
            "Usa nombres como alumnos_1.ods, alumnos_2.xls, etc."
        )

    resultado = {}
    profesor_actual = None
    alumnos_actuales = []

    def cerrar_bloque():
        nonlocal profesor_actual, alumnos_actuales
        if profesor_actual and alumnos_actuales:
            resultado[(seccion_forzada, profesor_actual)] = alumnos_actuales[:]
        profesor_actual = None
        alumnos_actuales = []

    for _, row in df_raw.iterrows():
        if es_fila_vacia(row):
            cerrar_bloque()
            continue

        primera = limpiar_texto(row.iloc[0] if len(row) > 0 else "")
        if not primera:
            continue

        prof_detectado = detectar_profesor_en_fila(primera)

        if prof_detectado:
            cerrar_bloque()
            profesor_actual = prof_detectado
            alumnos_actuales = []
            continue

        if profesor_actual:
            nombre = normalizar_nombre_alumno(primera)
            if nombre:
                alumnos_actuales.append(nombre)

    cerrar_bloque()
    return resultado

def cargar_alumnos_desde_carpeta(carpeta_curso, df_eventos):
    alumnos_por_seccion_profesor = {}

    # Prioridad 1: integrantes.csv
    ruta_csv = os.path.join(carpeta_curso, "integrantes.csv")
    if os.path.exists(ruta_csv):
        try:
            alumnos_por_seccion_profesor = cargar_alumnos_desde_integrantes_csv(ruta_csv, df_eventos)
            if alumnos_por_seccion_profesor:
                print(f"✅ Leído integrantes.csv en {carpeta_curso}")
                return alumnos_por_seccion_profesor
        except Exception as e:
            print(f"⚠️ No pude procesar integrantes.csv: {e}")

    # Prioridad 2: alumnos_1.ods/xls/xlsx
    if not os.path.exists(carpeta_curso):
        return alumnos_por_seccion_profesor

    archivos = []
    for nombre in os.listdir(carpeta_curso):
        lower = nombre.lower()
        if lower.startswith("alumnos_") and (lower.endswith(".ods") or lower.endswith(".xls") or lower.endswith(".xlsx")):
            archivos.append(os.path.join(carpeta_curso, nombre))

    archivos = sorted(archivos)

    for ruta in archivos:
        try:
            parciales = leer_alumnos_desde_archivo_por_bloques(ruta)
            for clave, alumnos in parciales.items():
                alumnos_por_seccion_profesor[clave] = alumnos
        except Exception as e:
            print(f"⚠️ No pude leer {ruta}: {e}")

    return alumnos_por_seccion_profesor

# ============================================================
# CONSTRUCCIÓN DE TABLAS DE ASISTENCIA
# ============================================================

def construir_bloques_asistencia(df_eventos, alumnos_por_seccion_profesor, curso_label):
    bloques = []

    if df_eventos.empty:
        return bloques

    agrupado = df_eventos.groupby(["sección", "actividad", "profesor"], dropna=False)

    for (seccion, actividad, profesor), sub in agrupado:
        sub = sub.sort_values("fecha").copy()

        etiquetas = []

        for _, r in sub.iterrows():
            n = int(r["n_clase"])
            ftxt = pd.to_datetime(r["fecha"]).strftime("%d/%m")
            etiquetas.append(f"{actividad} {n}\n{ftxt}")

        alumnos = alumnos_por_seccion_profesor.get((seccion, profesor), [])

        bloques.append({
            "curso": curso_label,
            "sección": seccion,
            "actividad": actividad,
            "profesor": profesor,
            "alumnos": alumnos,
            "etiquetas_columnas": etiquetas,
        })

    bloques = sorted(
        bloques,
        key=lambda x: (x["sección"], x["actividad"], x["profesor"])
    )

    return bloques

# ============================================================
# EXCEL
# ============================================================

def ancho_columna_fecha(n_fechas):
    if n_fechas <= 8:
        return 14
    if n_fechas <= 12:
        return 11
    if n_fechas <= 16:
        return 9
    return 8

def font_size_fecha(n_fechas):
    if n_fechas <= 8:
        return 10
    if n_fechas <= 12:
        return 9
    if n_fechas <= 16:
        return 8
    return 7

def alto_fila_header(n_fechas):
    if n_fechas <= 10:
        return 34
    if n_fechas <= 16:
        return 30
    return 26

def aplicar_estilo_hoja_asistencia(ws, n_alumnos, n_fechas):
    thin = Side(style="thin", color="000000")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    for fila in range(1, 5):
        for col in range(1, n_fechas + 2):
            ws.cell(row=fila, column=col).alignment = Alignment(vertical="center")

    fila_inicio = 5
    fila_fin = 5 + max(n_alumnos, 1)
    col_fin = n_fechas + 1

    for r in range(fila_inicio, fila_fin + 1):
        for c in range(1, col_fin + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = borde
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A5"].fill = PatternFill("solid", fgColor="D9EAF7")
    for c in range(2, col_fin + 1):
        ws.cell(row=5, column=c).fill = PatternFill("solid", fgColor="FCE4D6")
        ws.cell(row=5, column=c).font = Font(bold=True, size=font_size_fecha(n_fechas))

    ws["A5"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 28
    ancho_fechas = ancho_columna_fecha(n_fechas)
    for c in range(2, col_fin + 1):
        ws.column_dimensions[get_column_letter(c)].width = ancho_fechas

    ws.row_dimensions[5].height = alto_fila_header(n_fechas)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins = PageMargins(left=0.20, right=0.20, top=0.35, bottom=0.35, header=0.15, footer=0.15)

    ws.print_title_rows = "1:5"

def escribir_bloque_en_hoja(ws, bloque):
    curso = bloque["curso"]
    seccion = bloque["sección"]
    actividad = bloque["actividad"]
    profesor = bloque["profesor"]
    alumnos = bloque["alumnos"]
    etiquetas = bloque["etiquetas_columnas"]

    ws["A1"] = f"Profesor: {profesor}"
    ws["A2"] = f"Curso: {curso}"
    ws["A3"] = f"Sección: {seccion}"
    ws["A4"] = f"Actividad: {actividad}"

    for c in ["A1", "A2", "A3", "A4"]:
        ws[c].font = Font(bold=True, size=12)

    ws["A5"] = "Nombre del estudiante"

    for idx, etiqueta in enumerate(etiquetas, start=2):
        ws.cell(row=5, column=idx, value=etiqueta)

    if alumnos:
        for i, nombre in enumerate(alumnos, start=6):
            ws.cell(row=i, column=1, value=nombre)
            for j in range(2, len(etiquetas) + 2):
                ws.cell(row=i, column=j, value="")
    else:
        ws.cell(row=6, column=1, value="(Pegar aquí lista de estudiantes)")
        for j in range(2, len(etiquetas) + 2):
            ws.cell(row=6, column=j, value="")

    aplicar_estilo_hoja_asistencia(ws, max(len(alumnos), 1), len(etiquetas))

def generar_excel_asistencia(bloques, path_salida_excel):
    wb = Workbook()
    ws0 = wb.active
    wb.remove(ws0)

    if not bloques:
        ws = wb.create_sheet("Sin_datos")
        ws["A1"] = "No hay datos para generar asistencia."
    else:
        nombres_usados = set()

        for bloque in bloques:
            base = f"{bloque['profesor']}_{bloque['sección']}_{bloque['actividad']}"
            nombre = sheet_name_seguro(base)

            original = nombre
            k = 2
            while nombre in nombres_usados:
                suf = f"_{k}"
                nombre = sheet_name_seguro(original[:31 - len(suf)] + suf)
                k += 1

            nombres_usados.add(nombre)

            ws = wb.create_sheet(nombre)
            escribir_bloque_en_hoja(ws, bloque)

    asegurar_directorio(os.path.dirname(path_salida_excel))
    wb.save(path_salida_excel)

# ============================================================
# PDF
# ============================================================

def exportar_pdf_desde_excel(path_excel, output_dir_pdf):
    os.makedirs(output_dir_pdf, exist_ok=True)

    candidatos = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        os.path.expanduser("~/Applications/LibreOffice.app/Contents/MacOS/soffice"),
    ]

    ejecutable = None
    for ruta in candidatos:
        if ruta and os.path.exists(ruta):
            ejecutable = ruta
            break

    if ejecutable is None:
        print("⚠️ No se encontró LibreOffice/soffice.")
        return False

    try:
        subprocess.run(
            [
                ejecutable,
                "--headless",
                "--convert-to", "pdf",
                "--outdir", output_dir_pdf,
                path_excel,
            ],
            check=True,
        )
        print(f"✅ PDF exportado correctamente desde {path_excel}")
        return True
    except subprocess.CalledProcessError as e:
        print(f"⚠️ No pude exportar a PDF con LibreOffice: {path_excel}")
        print(e)
        return False

# ============================================================
# CURSO
# ============================================================

def procesar_curso(curso_key, info):
    carpeta = info["carpeta"]
    curso_label = info["label"]

    carpeta_curso = os.path.join(DATA_DIR, carpeta)
    excel_cal = os.path.join(carpeta_curso, "calendario.xlsx")

    if not os.path.exists(excel_cal):
        print(f"⚠️ [{curso_key}] No existe calendario.xlsx")
        return

    print(f"\n=== Procesando {curso_label} ===")

    df_cal = cargar_calendario_excel(excel_cal)
    if df_cal.empty:
        print(f"⚠️ [{curso_key}] Calendario vacío o no legible")
        return

    df_eventos = extraer_eventos_asistencia(df_cal)
    if df_eventos.empty:
        print(f"⚠️ [{curso_key}] No encontré Seminarios/Laboratorios")
        return

    alumnos_por_seccion_profesor = cargar_alumnos_desde_carpeta(carpeta_curso, df_eventos)

    print("Profes/Secciones con alumnos detectados:")
    for clave in sorted(alumnos_por_seccion_profesor.keys()):
        print("  ", clave, "->", len(alumnos_por_seccion_profesor[clave]), "alumnos")

    bloques = construir_bloques_asistencia(
        df_eventos=df_eventos,
        alumnos_por_seccion_profesor=alumnos_por_seccion_profesor,
        curso_label=curso_label
    )

    salida_dir = os.path.join(carpeta_curso, "asistencia")
    asegurar_directorio(salida_dir)

    path_excel = os.path.join(salida_dir, f"asistencia_{slugify(curso_label)}.xlsx")

    generar_excel_asistencia(bloques, path_excel)
    print(f"✅ Excel creado: {path_excel}")

    ok_pdf = exportar_pdf_desde_excel(path_excel, salida_dir)
    if ok_pdf:
        pdf_name = os.path.splitext(os.path.basename(path_excel))[0] + ".pdf"
        print(f"✅ PDF creado: {os.path.join(salida_dir, pdf_name)}")

# ============================================================
# MAIN
# ============================================================

def main():
    for curso_key, info in CURSOS.items():
        try:
            procesar_curso(curso_key, info)
        except Exception as e:
            print(f"❌ Error en {curso_key}: {e}")

if __name__ == "__main__":
    main()