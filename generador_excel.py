import os
import yaml
import pandas as pd
from datetime import time
import holidays

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ORDEN_DIAS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
DIAS_ES = {
    "Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
    "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"
}
DIAS_ES_POR_WEEKDAY = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"}

COLORES_ACTIVIDAD = {
    "Clase teórica": "DCEBFA",
    "Seminario": "D9EAD3",
    "Laboratorio": "FCE5CD",
    "Trabajo autónomo": "EADCF8",
    "Sin clases (Feriado)": "F4CCCC",
    "Sin clases (Pausa académica)": "F4CCCC",
    "Examen": "D9D9D9",
}

def cargar_config(path="calendario_config.yml"):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

def hhmm_a_time(s: str) -> time:
    h, m = s.split(":")
    return time(int(h), int(m))

def rango_horario_str(inicio: str, fin: str) -> str:
    return f"{inicio}–{fin}"

def fecha_de_dia_en_semana(lunes_semana: pd.Timestamp, dia_semana: str) -> pd.Timestamp:
    idx = ORDEN_DIAS.index(dia_semana)
    return pd.Timestamp(lunes_semana.date()) + pd.Timedelta(days=idx)

def normalizar_profesores(valor):
    """
    Acepta:
    - "JCS, TY, AR"
    - ["JCS", "TY", "AR"]
    - None
    y devuelve siempre un string: "JCS, TY, AR"
    """
    if valor is None:
        return ""

    if isinstance(valor, list):
        return ", ".join([str(x).strip() for x in valor if str(x).strip()])

    return str(valor).strip()

def construir_df(config: dict) -> pd.DataFrame:
    cfg = config["calendario"]

    inicio = pd.Timestamp(cfg["periodo"]["fecha_inicio"])
    nsem = int(cfg["periodo"]["numero_semanas"])

    filas = []

    for seccion, scfg in cfg["secciones"].items():
        for semana in range(1, nsem + 1):
            lunes = inicio + pd.Timedelta(days=7 * (semana - 1))

            for clave, nombre in [("teorica", "Clase teórica"), ("seminario", "Seminario")]:
                if clave not in scfg:
                    continue

                dia_semana = scfg[clave]["dia_semana"]
                fecha = fecha_de_dia_en_semana(lunes, dia_semana)

                filas.append({
                    "semana": semana,
                    "fecha": fecha.date(),
                    "día": DIAS_ES[dia_semana],
                    "horario": rango_horario_str(scfg[clave]["inicio"], scfg[clave]["fin"]),
                    "sección": seccion,
                    "actividad": nombre,
                    "tema": "",
                    "evaluación": "",
                    "profesores": "",
                    "observaciones": ""
                })

            # if cfg.get("laboratorios", {}).get("frecuencia") == "cada_2_semanas":
            #     if cfg.get("laboratorios", {}).get("semanas") == "pares" and (semana % 2 == 0) and ("lab" in scfg):
            #         dia_semana = scfg["lab"]["dia_semana"]
            #         fecha = fecha_de_dia_en_semana(lunes, dia_semana)

            #         filas.append({
            #             "semana": semana,
            #             "fecha": fecha.date(),
            #             "día": DIAS_ES[dia_semana],
            #             "horario": rango_horario_str(scfg["lab"]["inicio"], scfg["lab"]["fin"]),
            #             "sección": seccion,
            #             "actividad": "Laboratorio",
            #             "tema": "",
            #             "evaluación": "",
            #             "profesores": "",
            #             "observaciones": "Cada 2 semanas"
            #         })
            
            if "lab" in scfg:
                dia_semana = scfg["lab"]["dia_semana"]
                fecha = fecha_de_dia_en_semana(lunes, dia_semana)

                filas.append({
                    "semana": semana,
                    "fecha": fecha.date(),
                    "día": DIAS_ES[dia_semana],
                    "horario": rango_horario_str(scfg["lab"]["inicio"], scfg["lab"]["fin"]),
                    "sección": seccion,
                    "actividad": "Laboratorio",
                    "tema": "",
                    "evaluación": "",
                    "profesores": "",
                    "observaciones": ""
                })

    df = pd.DataFrame(filas)

    # temas = cfg.get("temas_por_semana", {}) or {}
    # if temas:
    #     temas_norm = {int(k): str(v) for k, v in temas.items()}
    #     df.loc[df["semana"].isin(list(temas_norm.keys())), "tema"] = df["semana"].map(temas_norm).fillna("")

    # ------------------------------------------------------------
    # Temas por tipo de actividad
    # ------------------------------------------------------------
    temas_teoricos = cfg.get("temas_teoricos", {}) or {}
    temas_seminarios = cfg.get("temas_seminarios", {}) or {}
    temas_laboratorios = cfg.get("temas_laboratorios", {}) or {}

    # Compatibilidad antigua: si no existen las nuevas listas, usar temas_por_semana
    temas_por_semana = cfg.get("temas_por_semana", {}) or {}
    if temas_por_semana:
        temas_por_semana = {int(k): str(v) for k, v in temas_por_semana.items()}
        if not temas_teoricos:
            temas_teoricos = temas_por_semana.copy()
        if not temas_seminarios:
            temas_seminarios = temas_por_semana.copy()
        if not temas_laboratorios:
            temas_laboratorios = temas_por_semana.copy()

    if temas_teoricos:
        temas_teoricos = {int(k): str(v) for k, v in temas_teoricos.items()}
        mask = (df["actividad"] == "Clase teórica") & (df["semana"].isin(list(temas_teoricos.keys())))
        df.loc[mask, "tema"] = df.loc[mask, "semana"].map(temas_teoricos).fillna("")

    if temas_seminarios:
        temas_seminarios = {int(k): str(v) for k, v in temas_seminarios.items()}
        mask = (df["actividad"] == "Seminario") & (df["semana"].isin(list(temas_seminarios.keys())))
        df.loc[mask, "tema"] = df.loc[mask, "semana"].map(temas_seminarios).fillna("")

    if temas_laboratorios:
        temas_laboratorios = {int(k): str(v) for k, v in temas_laboratorios.items()}
        mask = (df["actividad"] == "Laboratorio") & (df["semana"].isin(list(temas_laboratorios.keys())))
        df.loc[mask, "tema"] = df.loc[mask, "semana"].map(temas_laboratorios).fillna("")


    # # ------------------------------------------------------------
    # # Excepciones de tema
    # # ------------------------------------------------------------
    # for ex in cfg.get("excepciones_tema", []) or []:
    #     mask = (
    #         (df["sección"] == ex.get("seccion", "")) &
    #         (df["semana"] == int(ex.get("semana"))) &
    #         (df["actividad"] == ex.get("actividad", ""))
    #     )
    #     df.loc[mask, "tema"] = str(ex.get("tema", "")).strip()

    # ------------------------------------------------------------
    # Excepciones de evento
    # Permite cambiar semana, fecha, día, horario, tema, actividad, etc.
    # ------------------------------------------------------------
    for ex in cfg.get("excepciones_evento", []) or []:
        seccion_ex = ex.get("seccion", "")
        semana_ex = int(ex.get("semana"))
        actividad_ex = ex.get("actividad", "")

        mask = (
            (df["sección"] == seccion_ex) &
            (df["semana"] == semana_ex) &
            (df["actividad"] == actividad_ex)
        )

        if not mask.any():
            continue

        # semana base a usar para recalcular fecha
        semana_destino = int(ex.get("semana_nueva", semana_ex))
        df.loc[mask, "semana"] = semana_destino

        lunes_semana_destino = inicio + pd.Timedelta(days=7 * (semana_destino - 1))

        # día de la semana original o nuevo
        if "dia_semana" in ex:
            dia_semana_usar = str(ex["dia_semana"]).strip()
        else:
            dia_actual_es = df.loc[mask, "día"].iloc[0]
            mapa_es_a_en = {v: k for k, v in DIAS_ES.items()}
            dia_semana_usar = mapa_es_a_en.get(dia_actual_es, "Monday")

        fecha_nueva = fecha_de_dia_en_semana(lunes_semana_destino, dia_semana_usar)
        df.loc[mask, "fecha"] = fecha_nueva.date()
        df.loc[mask, "día"] = DIAS_ES[dia_semana_usar]

        # horario
        if "inicio" in ex and "fin" in ex:
            df.loc[mask, "horario"] = rango_horario_str(str(ex["inicio"]).strip(), str(ex["fin"]).strip())

        # tema
        if "tema" in ex:
            df.loc[mask, "tema"] = str(ex.get("tema", "")).strip()

        # actividad nueva
        if "actividad_nueva" in ex:
            df.loc[mask, "actividad"] = str(ex.get("actividad_nueva", "")).strip()

        # profesores
        if "profesores" in ex:
            df.loc[mask, "profesores"] = str(ex.get("profesores", "")).strip()

        # observaciones
        if "observaciones" in ex:
            df.loc[mask, "observaciones"] = str(ex.get("observaciones", "")).strip()

        # evaluación
        if "evaluacion" in ex:
            df.loc[mask, "evaluación"] = str(ex.get("evaluacion", "")).strip()

    # for r in cfg.get("profesores_base", []) or []:
    #     secc = r.get("seccion")
    #     act = r.get("actividad")
    #     prof = r.get("profesores", "")
    #     mask = (df["sección"] == secc) & (df["actividad"] == act)
    #     df.loc[mask, "profesores"] = prof

    for r in cfg.get("profesores_base", []) or []:
        secc = r.get("seccion")
        act = r.get("actividad")
        prof = normalizar_profesores(r.get("profesores", ""))
        mask = (df["sección"] == secc) & (df["actividad"] == act)
        df.loc[mask, "profesores"] = prof
    
    df = aplicar_evaluaciones(df, cfg, inicio)

    # for ev in cfg.get("evaluaciones", []) or []:
    #     if ev.get("modo") == "por_filtro":
    #         mask = (
    #             (df["sección"] == ev.get("seccion")) &
    #             (df["semana"] == int(ev.get("semana"))) &
    #             (df["actividad"] == ev.get("actividad"))
    #         )
    #         df.loc[mask, "evaluación"] = ev.get("tipo", "")
    #         obs = str(ev.get("observaciones", "")).strip()
    #         if obs:
    #             df.loc[mask, "observaciones"] = df.loc[mask, "observaciones"].astype(str).str.strip()
    #             df.loc[mask, "observaciones"] = df.loc[mask, "observaciones"].apply(
    #                 lambda x: (x + " | " if x else "") + obs
    #             )
    
    # for ev in cfg.get("evaluaciones", []) or []:
    #     modo = str(ev.get("modo", "por_filtro")).strip()

    #     if modo == "por_filtro":
    #         mask = (
    #             (df["sección"] == ev.get("seccion")) &
    #             (df["semana"] == int(ev.get("semana"))) &
    #             (df["actividad"] == ev.get("actividad"))
    #         )

    #         df.loc[mask, "evaluación"] = ev.get("tipo", "")

    #         obs = str(ev.get("observaciones", "")).strip()
    #         if obs:
    #             df.loc[mask, "observaciones"] = df.loc[mask, "observaciones"].astype(str).str.strip()
    #             df.loc[mask, "observaciones"] = df.loc[mask, "observaciones"].apply(
    #                 lambda x: (x + " | " if x and x != "nan" else "") + obs
    #             )

    #     elif modo == "por_fecha":
    #         seccion_ev = str(ev.get("seccion", "")).strip()
    #         fecha_ev = pd.Timestamp(ev.get("fecha")).date()
    #         tipo_ev = str(ev.get("tipo", "")).strip()
    #         obs = str(ev.get("observaciones", "")).strip()

    #         # buscar si ya existe un evento ese día para esa sección
    #         mask = (
    #             (df["sección"] == seccion_ev) &
    #             (df["fecha"] == fecha_ev)
    #         )

    #         if mask.any():
    #             # si ya existe una fila ese día, marcarla como evaluación
    #             df.loc[mask, "evaluación"] = tipo_ev

    #             if obs:
    #                 df.loc[mask, "observaciones"] = df.loc[mask, "observaciones"].astype(str).str.strip()
    #                 df.loc[mask, "observaciones"] = df.loc[mask, "observaciones"].apply(
    #                     lambda x: (x + " | " if x and x != "nan" else "") + obs
    #                 )
    #         else:
    #             # si no existe una fila ese día, crear una nueva
    #             fecha_ts = pd.Timestamp(fecha_ev)
    #             semana_ev = int(((fecha_ts - inicio).days) // 7) + 1

    #             nuevos = [{
    #                 "semana": semana_ev,
    #                 "fecha": fecha_ev,
    #                 "día": DIAS_ES_POR_WEEKDAY[fecha_ts.weekday()],
    #                 "horario": "",
    #                 "sección": seccion_ev,
    #                 "actividad": "Examen" if tipo_ev.lower() == "examen" else "Clase teórica",
    #                 "tema": obs,
    #                 "evaluación": tipo_ev,
    #                 "profesores": "",
    #                 "observaciones": obs
    #             }]

    #             df = pd.concat([df, pd.DataFrame(nuevos)], ignore_index=True)

    fer = cfg.get("feriados", {}) or {}
    feriados_map = {}

    if fer.get("usar_automaticos_chile", False):
        fmin = pd.Timestamp(df["fecha"].min()).date()
        fmax = pd.Timestamp(df["fecha"].max()).date()
        years = list(range(fmin.year, fmax.year + 1))
        cl = holidays.country_holidays("CL", years=years)
        for d, nombre in cl.items():
            if fmin <= d <= fmax:
                feriados_map[pd.Timestamp(d).date()] = str(nombre)

    for fm in fer.get("manuales", []) or []:
        feriados_map[pd.Timestamp(fm["fecha"]).date()] = str(fm.get("nombre", "Feriado"))

    if feriados_map:
        mask = df["fecha"].isin(list(feriados_map.keys()))
        df.loc[mask, "actividad"] = "Sin clases (Feriado)"
        df.loc[mask, "observaciones"] = df.loc[mask, "fecha"].map(feriados_map)

    for p in cfg.get("pausas_academicas", []) or []:
        ini = pd.Timestamp(p["inicio"]).date()
        fin = pd.Timestamp(p["fin"]).date()
        etiqueta = str(p.get("etiqueta", "Pausa académica")).strip()

        mask = (df["fecha"] >= ini) & (df["fecha"] <= fin)
        df.loc[mask, "actividad"] = "Sin clases (Pausa académica)"
        df.loc[mask, "observaciones"] = etiqueta

    semanas_auto = set(cfg.get("semanas_trabajo_autonomo", []) or [])
    if semanas_auto:
        mask = df["semana"].isin(list(semanas_auto))
        df.loc[mask, "actividad"] = "Trabajo autónomo"
        df.loc[mask, "observaciones"] = "No hay clases (trabajo autónomo)."

    ex = cfg.get("examenes", {}) or {}
    semanas_ex = set(ex.get("semanas_examenes", []) or [])
    ref = pd.Timestamp(cfg["periodo"]["fecha_inicio"])

    if semanas_ex:
        df = df[~df["semana"].isin(list(semanas_ex))].copy()

        nuevos_ex = []
        for e in ex.get("eventos", []) or []:
            fecha_ts = pd.Timestamp(e["fecha"])
            semana_ev = int((fecha_ts - ref).days // 7) + 1
            nuevos_ex.append({
                "semana": semana_ev,
                "fecha": fecha_ts.date(),
                "día": DIAS_ES_POR_WEEKDAY[fecha_ts.weekday()],
                "horario": rango_horario_str(e["inicio"], e["fin"]),
                "sección": e["seccion"],
                "actividad": e.get("actividad", "Examen"),
                "tema": e.get("tema", ""),
                "evaluación": e.get("evaluacion", ""),
                "profesores": e.get("profesores", ""),
                "observaciones": e.get("observaciones", "")
            })

        if nuevos_ex:
            df = pd.concat([df, pd.DataFrame(nuevos_ex)], ignore_index=True)

    bloques = cfg.get("bloques", {}) or {}
    defin = (bloques.get("definicion", {}) or {})
    protegidos = bloques.get("protegidos", []) or []

    def parse_horario(h: str):
        a, b = h.split("–")
        ha, ma = map(int, a.split(":"))
        hb, mb = map(int, b.split(":"))
        return time(ha, ma), time(hb, mb)

    def overlap(a1, a2, b1, b2):
        return (a1 < b2) and (b1 < a2)

    if defin and protegidos:
        for bp in protegidos:
            fecha_bp = pd.Timestamp(bp["fecha"]).date()
            bloque_id = str(bp["bloque"])
            if bloque_id not in defin:
                continue

            b_ini = hhmm_a_time(defin[bloque_id]["inicio"])
            b_fin = hhmm_a_time(defin[bloque_id]["fin"])

            mask_fecha = (df["fecha"] == fecha_bp)
            idxs = df.index[mask_fecha].tolist()

            for i in idxs:
                h = str(df.at[i, "horario"] or "").strip()
                if "–" not in h:
                    continue
                e_ini, e_fin = parse_horario(h)
                if overlap(e_ini, e_fin, b_ini, b_fin):
                    obs = str(df.at[i, "observaciones"] or "").strip()
                    tag = f"CONFLICTO: Bloque protegido {bloque_id}"
                    df.at[i, "observaciones"] = (obs + " | " if obs else "") + tag

    def hora_inicio(h):
        if isinstance(h, str) and "–" in h:
            return h.split("–")[0]
        return "00:00"

    df["_inicio_dt"] = pd.to_datetime(
        df["fecha"].astype(str) + " " + df["horario"].fillna("").apply(hora_inicio),
        errors="coerce"
    )
    df = df.sort_values(["sección", "fecha", "_inicio_dt"], na_position="last").drop(columns=["_inicio_dt"])

    cols = [
        "semana", "fecha", "día", "horario", "sección",
        "actividad", "tema", "evaluación", "profesores", "observaciones"
    ]
    return df[cols]

def combinar_observaciones(obs_actual, obs_nueva):
    obs_actual = str(obs_actual or "").strip()
    obs_nueva = str(obs_nueva or "").strip()

    if not obs_nueva:
        return obs_actual
    if not obs_actual or obs_actual == "nan":
        return obs_nueva
    return f"{obs_actual} | {obs_nueva}"


def calcular_semana_desde_fecha(fecha, inicio_semestre, nsem):
    fecha_ts = pd.Timestamp(fecha)
    inicio_ts = pd.Timestamp(inicio_semestre)

    semana = int(((fecha_ts.normalize() - inicio_ts.normalize()).days) // 7) + 1

    if semana < 1:
        semana = 1
    if semana > nsem:
        semana = nsem

    return semana


def aplicar_evaluaciones(df, cfg, inicio):
    """
    Soporta:
    - modo: por_filtro
    - modo: por_fecha
        * si viene inicio+fin => horario específico
        * si no => todo el día
    """
    evaluaciones = cfg.get("evaluaciones", []) or []
    nsem = int(cfg["periodo"]["numero_semanas"])

    filas_nuevas = []

    for ev in evaluaciones:
        modo = str(ev.get("modo", "por_filtro")).strip()

        # =====================================================
        # 1) EVALUACIÓN SOBRE UNA CLASE YA EXISTENTE
        # =====================================================
        if modo == "por_filtro":
            seccion = str(ev.get("seccion", "")).strip()
            semana = int(ev.get("semana"))
            actividad = str(ev.get("actividad", "")).strip()
            tipo = str(ev.get("tipo", "")).strip()
            obs = str(ev.get("observaciones", "")).strip()

            mask = (
                (df["sección"] == seccion) &
                (df["semana"] == semana) &
                (df["actividad"] == actividad)
            )

            df.loc[mask, "evaluación"] = tipo

            if obs:
                df.loc[mask, "observaciones"] = df.loc[mask, "observaciones"].apply(
                    lambda x: combinar_observaciones(x, obs)
                )

        # =====================================================
        # 2) EVALUACIÓN EXTRA POR FECHA
        # =====================================================
        elif modo == "por_fecha":
            seccion = str(ev.get("seccion", "")).strip()
            fecha_raw = ev.get("fecha", None)
            tipo = str(ev.get("tipo", "")).strip()

            if not fecha_raw or not tipo or not seccion:
                continue

            fecha_ts = pd.Timestamp(fecha_raw)
            fecha_date = fecha_ts.date()

            inicio_ev = str(ev.get("inicio", "")).strip()
            fin_ev = str(ev.get("fin", "")).strip()
            actividad_ev = str(ev.get("actividad", "")).strip()
            tema_ev = str(ev.get("tema", "")).strip()
            obs = str(ev.get("observaciones", "")).strip()
            profesores_ev = normalizar_profesores(ev.get("profesores", ""))

            if not actividad_ev:
                actividad_ev = "Examen" if tipo.lower() == "examen" else "Evaluación especial"

            horario_ev = ""
            if inicio_ev and fin_ev:
                horario_ev = rango_horario_str(inicio_ev, fin_ev)

            semana_ev = calcular_semana_desde_fecha(fecha_ts, inicio, nsem)

            filas_nuevas.append({
                "semana": semana_ev,
                "fecha": fecha_date,
                "día": DIAS_ES_POR_WEEKDAY[fecha_ts.weekday()],
                "horario": horario_ev,
                "sección": seccion,
                "actividad": actividad_ev,
                "tema": tema_ev,
                "evaluación": tipo,
                "profesores": profesores_ev,
                "observaciones": obs
            })

    if filas_nuevas:
        df_extra = pd.DataFrame(filas_nuevas)

        for col in df.columns:
            if col not in df_extra.columns:
                df_extra[col] = ""

        for col in df_extra.columns:
            if col not in df.columns:
                df[col] = ""

        df = pd.concat([df, df_extra[df.columns]], ignore_index=True)

    return df


def exportar_excel(df: pd.DataFrame, path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
        df.to_excel(writer, index=False, sheet_name="Calendario")
        ws = writer.sheets["Calendario"]

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        fill_header = PatternFill(fill_type="solid", fgColor="1F4E78")
        font_header = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [c.value for c in ws[1]]
        idx_fecha = headers.index("fecha") + 1
        idx_horario = headers.index("horario") + 1
        idx_actividad = headers.index("actividad") + 1
        idx_evaluacion = headers.index("evaluación") + 1

        for r in range(2, ws.max_row + 1):
            c_fecha = ws.cell(row=r, column=idx_fecha)
            c_horario = ws.cell(row=r, column=idx_horario)
            c_actividad = ws.cell(row=r, column=idx_actividad)
            c_eval = ws.cell(row=r, column=idx_evaluacion)

            if isinstance(c_fecha.value, str) and c_fecha.value.strip():
                try:
                    c_fecha.value = pd.to_datetime(c_fecha.value).date()
                except Exception:
                    pass
            c_fecha.number_format = "DD/MM/YYYY"
            c_fecha.alignment = Alignment(horizontal="center")

            if c_horario.value is None:
                c_horario.value = ""
            c_horario.number_format = "@"
            c_horario.alignment = Alignment(horizontal="center")

            actividad = str(c_actividad.value or "").strip()
            color = COLORES_ACTIVIDAD.get(actividad)
            if color:
                fill_row = PatternFill(fill_type="solid", fgColor=color)
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill_row

            if c_eval.value not in [None, ""]:
                c_eval.font = Font(bold=True)
                c_eval.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)

            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 45)

def main():
    cursos = {
        "fokito": {
            "config_path": os.path.join("config", "calendario_fokito.yml"),
            "output_path": os.path.join("data", "fokito", "calendario.xlsx"),
        },
        "tecnologia_medica": {
            "config_path": os.path.join("config", "calendario_tecnologia_medica.yml"),
            "output_path": os.path.join("data", "tecnologia_medica", "calendario.xlsx"),
        },
        "medicina": {
            "config_path": os.path.join("config", "calendario_medicina.yml"),
            "output_path": os.path.join("data", "medicina", "calendario.xlsx"),
        },
        
        "enobnu": {
            "config_path": os.path.join("config", "calendario_enobnu.yml"),
            "output_path": os.path.join("data", "enobnu", "calendario.xlsx"),
        },
    }

    for curso, info in cursos.items():
        print(curso)
        config_path = info["config_path"]
        output_path = info["output_path"]

        if not os.path.exists(config_path):
            print(f"⚠️  Saltando {curso}: no existe {config_path}")
            continue

        config = cargar_config(config_path)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        df = construir_df(config)
        exportar_excel(df, output_path)

        print(f"OK: [{curso}] Excel generado en {output_path} con {len(df)} filas.")

if __name__ == "__main__":
    main()