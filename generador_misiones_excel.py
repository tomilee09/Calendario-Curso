import os
import yaml
import pandas as pd

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# YAML / HELPERS
# ============================================================
def cargar_yaml(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def split_profes(valor):
    """
    Acepta:
    - None
    - "A, B, C"
    - ["A", "B", "C"]
    y preserva duplicados si vienen en lista/string.
    """
    if valor is None:
        return []

    try:
        if pd.isna(valor):
            return []
    except Exception:
        pass

    if isinstance(valor, (list, tuple)):
        return [str(x).strip() for x in valor if str(x).strip()]

    texto = str(valor).strip()
    if not texto:
        return []

    return [x.strip() for x in texto.split(",") if x.strip()]


def normalizar_lista_profes(valor):
    return split_profes(valor)


def normalizar_profes_str(valor):
    return ", ".join(split_profes(valor))


def unicos(lista):
    salida = []
    for x in lista:
        sx = str(x).strip()
        if sx and sx not in salida:
            salida.append(sx)
    return salida


def a_fecha(x):
    ts = pd.to_datetime(x, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.date()


def construir_horas_seminario_map(df_cal):
    df_horas = horas_seminario_por_profesor(df_cal)
    if df_horas.empty:
        return {}
    return dict(zip(df_horas["profesor"], df_horas["horas_seminario"]))


def score_balance(profesor, carga_actual, horas_map):
    horas = float(horas_map.get(profesor, 0.0))
    carga = float(carga_actual.get(profesor, 0.0))

    if horas <= 0:
        return -10**9

    return horas / (carga + 1.0)


def elegir_balanceado(pool, carga_actual, horas_map, cantidad=1, excluir=None):
    if excluir is None:
        excluir = set()
    else:
        excluir = set(excluir)

    candidatos = [p for p in unicos(pool) if p not in excluir]

    if not candidatos:
        return []

    candidatos = sorted(
        candidatos,
        key=lambda p: (
            score_balance(p, carga_actual, horas_map),
            float(horas_map.get(p, 0.0)),
            -float(carga_actual.get(p, 0.0)),
            p
        ),
        reverse=True
    )

    return candidatos[:cantidad]


def sumar_carga(carga_actual, responsables, peso=1.0):
    for p in split_profes(responsables):
        carga_actual[p] = float(carga_actual.get(p, 0.0)) + float(peso)


def construir_id_mision(row):
    """
    ID estable para preservar estado entre regeneraciones.
    """
    fecha_limite = pd.to_datetime(row.get("fecha_limite"), errors="coerce")
    fecha_evento = pd.to_datetime(row.get("fecha_evento"), errors="coerce")

    fecha_limite_str = fecha_limite.strftime("%Y-%m-%d") if pd.notna(fecha_limite) else ""
    fecha_evento_str = fecha_evento.strftime("%Y-%m-%d") if pd.notna(fecha_evento) else ""

    partes = [
        fecha_limite_str,
        fecha_evento_str,
        str(row.get("evento", "")).strip(),
        str(row.get("tipo_evento", "")).strip(),
        str(row.get("paso", "")).strip(),
        str(row.get("sección", "")).strip(),
        str(row.get("responsables", "")).strip(),
    ]
    return "||".join(partes)


def cargar_estados_previos(path_excel):
    """
    Lee el Excel previo y devuelve {id_mision: estado}
    """
    if not os.path.exists(path_excel):
        return {}

    try:
        df_prev = pd.read_excel(path_excel, sheet_name="Misiones")
    except Exception:
        return {}

    if df_prev.empty:
        return {}

    for c in ["fecha_limite", "fecha_evento"]:
        if c in df_prev.columns:
            df_prev[c] = pd.to_datetime(df_prev[c], errors="coerce")

    for c in ["evento", "tipo_evento", "paso", "sección", "responsables", "estado"]:
        if c in df_prev.columns:
            df_prev[c] = df_prev[c].fillna("").astype(str)
        else:
            df_prev[c] = ""

    estados = {}
    for _, row in df_prev.iterrows():
        id_mision = construir_id_mision(row)
        estado = str(row.get("estado", "")).strip() or "Pendiente"
        estados[id_mision] = estado

    return estados


def aplicar_estados_previos(df_mis, estados_previos):
    if df_mis.empty:
        return df_mis

    df = df_mis.copy()
    df["id_mision"] = df.apply(construir_id_mision, axis=1)
    df["estado"] = df["id_mision"].apply(lambda x: estados_previos.get(x, "Pendiente"))
    return df


# ============================================================
# LECTURA DE CONFIG DEL CURSO
# ============================================================
def obtener_profesores_base(config):
    """
    Devuelve dict:
    {(seccion, actividad): [prof1, prof2, ...]}
    """
    salida = {}
    for r in config.get("calendario", {}).get("profesores_base", []) or []:
        secc = str(r.get("seccion", "")).strip()
        act = str(r.get("actividad", "")).strip()
        profes = split_profes(r.get("profesores", []))
        salida[(secc, act)] = profes
    return salida


def obtener_coordinacion(config):
    """
    Lee:
    misiones:
      coordinacion:
        pec_por_seccion:
          "Sección 1": "CC"
        pcc_por_seccion:
          "Sección 1": "TY"
    """
    mis = config.get("misiones", {}) or {}
    coord = mis.get("coordinacion", {}) or {}

    pec_por_seccion = coord.get("pec_por_seccion", {}) or {}
    pcc_por_seccion = coord.get("pcc_por_seccion", {}) or {}

    return pec_por_seccion, pcc_por_seccion


def obtener_lista_seminario_seccion(config, seccion, mantener_duplicados=True):
    """
    Devuelve la lista de profesores de seminario de una sección.
    - Si mantener_duplicados=True, respeta repeticiones como ["SM", "XX", "XX"].
    - Si False, devuelve únicos preservando orden.
    """
    mapa = obtener_profesores_base(config)
    lista = mapa.get((seccion, "Seminario"), []).copy()

    if mantener_duplicados:
        return [str(x).strip() for x in lista if str(x).strip()]

    return unicos(lista)


# def obtener_pool_participantes_seccion(config, seccion):
#     """
#     Profesores participantes de la sección para misiones colectivas.
#     IMPORTANTE:
#     - NO excluye PEC ni PCC.
#     - Si alguien hace seminario, cuenta como participante aunque además sea PEC/PCC.
#     """
#     return obtener_lista_seminario_seccion(config, seccion, mantener_duplicados=False)


def obtener_pool_participantes_seccion(config, seccion):
    """
    Pool de profesores participantes de la sección para misiones colectivas.
    IMPORTANTE:
    - mantiene duplicados
    - si un profesor aparece dos veces en seminario, también aparece dos veces aquí
    """
    return obtener_lista_seminario_seccion(config, seccion, mantener_duplicados=True)


def obtener_pool_controles_seccion(config, seccion):
    """
    Pool ponderado para controles.
    Aquí sí se respetan duplicados del YAML, por ejemplo:
    ["SM", "XX", "XX"] -> XX tiene doble peso.
    """
    return obtener_lista_seminario_seccion(config, seccion, mantener_duplicados=True)




def obtener_pool_global_participantes(config):
    """
    Pool global ponderado para talleres.
    Junta todas las listas de seminario de todas las secciones,
    respetando duplicados dentro de cada sección.
    """
    secciones = list((config.get("calendario", {}).get("secciones", {}) or {}).keys())
    salida = []

    for seccion in secciones:
        salida.extend(obtener_lista_seminario_seccion(config, seccion, mantener_duplicados=True))

    return [str(x).strip() for x in salida if str(x).strip()]


def obtener_pool_laboratorio_seccion(config, seccion):
    mapa = obtener_profesores_base(config)
    return mapa.get((seccion, "Laboratorio"), []).copy()


# ============================================================
# POOLS / OVERRIDES
# ============================================================
def construir_pools_por_seccion(config):
    """
    Pool cíclico automático desde profesores_base / Seminario.
    Respeta duplicados para ponderar carga.
    Ejemplo:
      ["SM", "XX", "XX"] -> slots A=SM, B=XX, C=XX
    """
    cfg_cal = config.get("calendario", {}) or {}
    base = cfg_cal.get("profesores_base", []) or []

    pools = {}
    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    for r in base:
        seccion = str(r.get("seccion", "")).strip()
        actividad = str(r.get("actividad", "")).strip()
        profesores = split_profes(r.get("profesores", ""))

        if not seccion or actividad != "Seminario":
            continue

        pool = []
        for i, prof in enumerate(profesores):
            if i >= len(letras):
                break
            pool.append({
                "slot": letras[i],
                "profesor": prof
            })

        pools[seccion] = pool

    return pools


def construir_overrides(cfg_mis):
    """
    Índice por:
    (tipo_evento, seccion, fecha_evento)
    """
    overrides = {}
    for r in cfg_mis.get("overrides_responsables", []) or []:
        tipo = str(r.get("tipo_evento", "")).strip()
        seccion = str(r.get("seccion", "")).strip()
        fecha = a_fecha(r.get("fecha_evento"))
        if not tipo or not seccion or not fecha:
            continue

        overrides[(tipo, seccion, fecha)] = r

    return overrides


# def buscar_override(config, tipo_evento, seccion, fecha_evento):
#     mis = config.get("misiones", {}) or {}
#     overrides = mis.get("overrides_responsables", []) or []

#     fecha_evento_str = pd.Timestamp(fecha_evento).strftime("%Y-%m-%d")

#     for ov in overrides:
#         tipo_ov = str(ov.get("tipo_evento", "")).strip().lower()
#         seccion_ov = str(ov.get("seccion", "")).strip()
#         fecha_ov = str(ov.get("fecha_evento", "")).strip()

#         if tipo_ov == str(tipo_evento).strip().lower() and seccion_ov == seccion and fecha_ov == fecha_evento_str:
#             return ov

#     return None



def buscar_override(config, tipo_evento, seccion, fecha_evento, paso=None):
    mis = config.get("misiones", {}) or {}
    overrides = mis.get("overrides_responsables", []) or []

    fecha_evento_str = pd.Timestamp(fecha_evento).strftime("%Y-%m-%d")

    for ov in overrides:
        tipo_ov = str(ov.get("tipo_evento", "")).strip().lower()
        seccion_ov = str(ov.get("seccion", "")).strip()
        fecha_ov = str(ov.get("fecha_evento", "")).strip()
        paso_ov = str(ov.get("paso", "")).strip()

        if tipo_ov != str(tipo_evento).strip().lower():
            continue
        if seccion_ov != seccion:
            continue
        if fecha_ov != fecha_evento_str:
            continue

        if paso is not None and paso_ov:
            if paso_ov != paso:
                continue

        return ov

    return None





def slot_a_indice(slot):
    slot = str(slot).strip().upper()
    mapa = {
        "A": 0, "B": 1, "C": 2, "D": 3,
        "E": 4, "F": 5, "G": 6, "H": 7
    }
    return mapa.get(slot, None)


def elegir_ciclico(pool, indice):
    if not pool:
        return []
    return [pool[indice % len(pool)]]


# def elegir_varios_ciclico_sin_repetir(pool, inicio, cantidad):
#     """
#     Elige 'cantidad' profesores recorriendo un pool ponderado en forma cíclica,
#     pero evitando repetir a la misma persona dentro de la misma asignación.
#     """
#     if not pool or cantidad <= 0:
#         return []

#     n = len(pool)
#     elegidos = []
#     vistos = set()
#     intentos = 0
#     idx = inicio

#     while len(elegidos) < cantidad and intentos < 10 * n:
#         prof = str(pool[idx % n]).strip()
#         if prof and prof not in vistos:
#             elegidos.append(prof)
#             vistos.add(prof)
#         idx += 1
#         intentos += 1

#     return elegidos


def elegir_varios_ciclico_sin_repetir(pool, inicio, cantidad):
    """
    Elige 'cantidad' profesores recorriendo un pool ponderado en forma cíclica,
    evitando repetir a la misma persona dentro de la misma asignación.

    IMPORTANTE:
    devuelve también el nuevo cursor REAL, según cuántas posiciones
    se recorrieron efectivamente en el pool.
    """
    if not pool or cantidad <= 0:
        return [], inicio

    n = len(pool)
    elegidos = []
    vistos = set()
    intentos = 0
    idx = inicio

    while len(elegidos) < cantidad and intentos < 10 * n:
        prof = str(pool[idx % n]).strip()
        if prof and prof not in vistos:
            elegidos.append(prof)
            vistos.add(prof)
        idx += 1
        intentos += 1

    return elegidos, idx

# ============================================================
# DETECCIÓN DE EVALUACIONES
# ============================================================
def nombre_evento_desde_ev(ev):
    """
    Define un nombre amigable del evento sin confundirlo con observaciones
    administrativas o feriados.
    """
    tipo = str(ev.get("tipo_evento", "")).strip()
    nombre_base = str(ev.get("nombre_evento", tipo)).strip()
    obs = str(ev.get("observaciones", "")).strip()

    obs_low = obs.lower()

    palabras_feriado = [
        "san pedro", "san pablo", "feriado", "pausa académica",
        "trabajo autónomo", "receso", "no hay clases"
    ]

    if obs and not any(p in obs_low for p in palabras_feriado):
        return obs

    return nombre_base


def detectar_evaluaciones(df_cal):
    df = df_cal.copy()

    if "fecha" in df.columns:
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")

    for c in ["evaluación", "actividad", "sección", "tema", "observaciones"]:
        if c in df.columns:
            df[c] = df[c].fillna("").astype(str)

    evs = []

    sub = df[df["evaluación"].str.strip() != ""].copy()

    for _, r in sub.iterrows():
        evs.append({
            "tipo_evento": str(r["evaluación"]).strip(),
            "nombre_evento": str(r["evaluación"]).strip(),
            "sección": str(r.get("sección", "")).strip(),
            "actividad": str(r.get("actividad", "")).strip(),
            "fecha_evento": pd.Timestamp(r["fecha"]).date(),
            "tema": str(r.get("tema", "")).strip(),
            "observaciones": str(r.get("observaciones", "")).strip(),
        })

    df_evs = pd.DataFrame(evs)

    if df_evs.empty:
        return df_evs

    df_evs = df_evs.sort_values(
        ["fecha_evento", "sección", "tipo_evento", "actividad", "nombre_evento"]
    ).reset_index(drop=True)

    return df_evs







import matplotlib.pyplot as plt


def horas_seminario_por_profesor(df_cal):
    df = df_cal.copy()

    for c in ["actividad", "profesores", "horario"]:
        if c not in df.columns:
            df[c] = ""

    df = df[df["actividad"].astype(str).str.strip() == "Seminario"].copy()

    filas = []
    for _, row in df.iterrows():
        horario = str(row.get("horario", "")).strip()
        profes = split_profes(row.get("profesores", ""))

        duracion = 0.0
        m = pd.Series([horario]).astype(str).str.replace("–", "-", regex=False).iloc[0]
        mm = __import__("re").match(r"^\s*(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})\s*$", m)
        if mm:
            t1 = pd.to_datetime(mm.group(1), format="%H:%M", errors="coerce")
            t2 = pd.to_datetime(mm.group(2), format="%H:%M", errors="coerce")
            if pd.notna(t1) and pd.notna(t2):
                duracion = (t2 - t1).total_seconds() / 3600.0

        for p in profes:
            filas.append({"profesor": p, "horas": duracion})

    if not filas:
        return pd.DataFrame(columns=["profesor", "horas_seminario"])

    out = pd.DataFrame(filas).groupby("profesor", as_index=False)["horas"].sum()
    out = out.rename(columns={"horas": "horas_seminario"})
    return out


# def misiones_seminario_por_profesor(df_mis):
#     if df_mis.empty:
#         return pd.DataFrame(columns=["profesor", "misiones_seminario"])

#     df = df_mis.copy()

#     # excluimos laboratorio
#     df = df[df["tipo_evento"].astype(str).str.strip() != "Laboratorio"].copy()

#     filas = []
#     for _, row in df.iterrows():
#         for p in split_profes(row.get("responsables", "")):
#             filas.append({"profesor": p, "misiones_seminario": 1})

#     if not filas:
#         return pd.DataFrame(columns=["profesor", "misiones_seminario"])

#     out = pd.DataFrame(filas).groupby("profesor", as_index=False)["misiones_seminario"].sum()
#     return out


def misiones_seminario_por_profesor(df_mis):
    if df_mis.empty:
        return pd.DataFrame(columns=["profesor", "misiones_seminario"])

    df = df_mis.copy()

    # solo misiones realmente repartibles respecto a horas de seminario
    pasos_validos = {
        "construir_control",
        "escanear",
        "corregir_y_notas",
        "revisar_tp",
        "construir_taller",
        "corregir_taller",
        "revisar_portafolio",
    }

    df = df[df["paso"].astype(str).str.strip().isin(pasos_validos)].copy()

    filas = []
    for _, row in df.iterrows():
        for p in split_profes(row.get("responsables", "")):
            filas.append({"profesor": p, "misiones_seminario": 1})

    if not filas:
        return pd.DataFrame(columns=["profesor", "misiones_seminario"])

    out = pd.DataFrame(filas).groupby("profesor", as_index=False)["misiones_seminario"].sum()
    return out


def graficar_equilibrio_curso(nombre_curso, df_cal, df_mis, carpeta_salida):
    df_horas = horas_seminario_por_profesor(df_cal)
    df_misiones_prof = misiones_seminario_por_profesor(df_mis)

    df_plot = pd.merge(df_horas, df_misiones_prof, on="profesor", how="outer").fillna(0)

    if df_plot.empty:
        print(f"⚠️ No hay datos para graficar equilibrio en {nombre_curso}")
        return

    df_plot["ratio_horas_por_mision"] = df_plot.apply(
        lambda r: (r["horas_seminario"] / r["misiones_seminario"]) if r["misiones_seminario"] > 0 else 0,
        axis=1
    )

    df_plot = df_plot.sort_values("ratio_horas_por_mision", ascending=False).reset_index(drop=True)

    plt.figure(figsize=(10, 5))
    plt.bar(df_plot["profesor"], df_plot["ratio_horas_por_mision"])
    plt.xticks(rotation=45, ha="right")
    plt.ylabel("Horas de seminario / misiones")
    plt.title(f"Equilibrio de carga — {nombre_curso}")
    plt.tight_layout()

    os.makedirs(carpeta_salida, exist_ok=True)
    path_png = os.path.join(carpeta_salida, "equilibrio_carga.png")
    plt.savefig(path_png, dpi=200)
    plt.close()

    print(f"✅ Gráfico de equilibrio guardado en: {path_png}")
    print(df_plot.to_string(index=False))






# ============================================================
# CONSTRUCCIÓN DE MISIONES
# ============================================================
def construir_misiones(config, df_cal):
    cfg_mis = config.get("misiones", {}) or {}
    protocolos = cfg_mis.get("protocolos", {}) or {}

    df_evs = detectar_evaluaciones(df_cal)

    if df_evs.empty:
        columnas = [
            "fecha_limite", "fecha_evento", "evento", "tipo_evento", "paso",
            "sección", "responsables", "detalle", "estado"
        ]
        return (
            pd.DataFrame(columns=columnas),
            pd.DataFrame(),
            pd.DataFrame()
        )

    df_evs["fecha_evento"] = pd.to_datetime(df_evs["fecha_evento"], errors="coerce")
    df_evs = df_evs.sort_values(
        ["fecha_evento", "sección", "tipo_evento", "nombre_evento"]
    ).reset_index(drop=True)

    filas = []
    filas_chequeo = []
    filas_pools = []

    contador_control_por_seccion = {}
    contador_lab_video_por_seccion = {}
    contador_taller_global = 0
    horas_map = construir_horas_seminario_map(df_cal)
    carga_actual = {}

    pec_por_seccion, pcc_por_seccion = obtener_coordinacion(config)
    pools_por_seccion = construir_pools_por_seccion(config)

    for seccion, pool in pools_por_seccion.items():
        for item in pool:
            filas_pools.append({
                "sección": seccion,
                "slot": item["slot"],
                "profesor": item["profesor"],
            })

    for _, ev in df_evs.iterrows():
        tipo = str(ev.get("tipo_evento", "")).strip()
        nombre = nombre_evento_desde_ev(ev)
        seccion = str(ev.get("sección", "")).strip()
        fecha_evento = pd.Timestamp(ev["fecha_evento"])
        tema = str(ev.get("tema", "")).strip()
        obs_cal = str(ev.get("observaciones", "")).strip()

        tipo_norm = tipo.lower()

        if "control" in tipo_norm:
            proto_nombre = "Control"
        elif "taller" in tipo_norm:
            proto_nombre = "Taller"
        elif "certamen" in tipo_norm or "prueba" in tipo_norm:
            proto_nombre = "Certamen"
        elif "portafolio" in tipo_norm:
            proto_nombre = "Portafolio"
        elif "trabajo práctico" in tipo_norm or tipo_norm == "tp":
            proto_nombre = "Trabajo práctico"
        elif "examen" in tipo_norm:
            proto_nombre = "Examen"
        elif "laboratorio" in tipo_norm or "informe laboratorio" in tipo_norm:
            proto_nombre = "Laboratorio"
        else:
            continue

        proto = protocolos.get(proto_nombre, {})
        if not proto:
            continue

        participantes = obtener_pool_participantes_seccion(config, seccion)
        pool_lab = obtener_pool_laboratorio_seccion(config, seccion)
        pec = str(pec_por_seccion.get(seccion, "")).strip()
        pcc = str(pcc_por_seccion.get(seccion, "")).strip()

        override = buscar_override(config, proto_nombre, seccion, fecha_evento.date())

        detalle_base_evento = []
        if obs_cal:
            detalle_base_evento.append(obs_cal)
        if tema:
            detalle_base_evento.append(f"Tema: {tema}")
        detalle_base_evento = " — ".join(detalle_base_evento)

        # ========================================================
        # CONTROL
        # pool cíclico ponderado por sección
        # ========================================================
        # if proto_nombre == "Control":
        #     idx = contador_control_por_seccion.get(seccion, 0)

        #     pool_control = obtener_pool_controles_seccion(config, seccion)
        #     pool_slots = pools_por_seccion.get(seccion, [])

        #     responsable_control = []
        #     slot_usado = ""

        #     if override and "responsables" in override:
        #         responsable_control = normalizar_lista_profes(override.get("responsables", []))
        #         slot_usado = "MANUAL"

        #     elif override and "slot" in override:
        #         slot_override = str(override.get("slot", "")).strip().upper()
        #         encontrado = False

        #         for item in pool_slots:
        #             if str(item.get("slot", "")).strip().upper() == slot_override:
        #                 responsable_control = [str(item.get("profesor", "")).strip()]
        #                 slot_usado = slot_override
        #                 encontrado = True
        #                 break

        #         if not encontrado and pool_control:
        #             responsable_control = [str(pool_control[idx % len(pool_control)]).strip()]
        #             slot_usado = f"AUTO-{(idx % len(pool_control)) + 1}"

        #     else:
        #         if pool_control:
        #             responsable_control = [str(pool_control[idx % len(pool_control)]).strip()]
        #             if pool_slots:
        #                 slot_usado = str(pool_slots[idx % len(pool_slots)].get("slot", "")).strip()
        #             else:
        #                 slot_usado = f"AUTO-{(idx % len(pool_control)) + 1}"

        #     contador_control_por_seccion[seccion] = idx + 1

        #     filas_chequeo.append({
        #         "sección": seccion,
        #         "fecha_evento": fecha_evento.date(),
        #         "evento": nombre,
        #         "observaciones": obs_cal,
        #         "slot_pool": slot_usado,
        #         "responsable_control": normalizar_profes_str(responsable_control),
        #     })

        #     for paso_key, paso in proto.items():
        #         offset = int(paso.get("offset_dias", 0))
        #         deadline = (fecha_evento + pd.Timedelta(days=offset)).date()

        #         detalle = str(paso.get("detalle", "")).strip()
        #         if detalle_base_evento:
        #             detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

        #         filas.append({
        #             "fecha_limite": deadline,
        #             "fecha_evento": fecha_evento.date(),
        #             "evento": nombre,
        #             "tipo_evento": proto_nombre,
        #             "paso": paso_key,
        #             "sección": seccion,
        #             "responsables": normalizar_profes_str(responsable_control),
        #             "detalle": detalle,
        #             "estado": "Pendiente",
        #         })
        #     continue
        
        if proto_nombre == "Control":
            pool_control = obtener_pool_controles_seccion(config, seccion)
            pool_slots = pools_por_seccion.get(seccion, [])

            responsable_control = []
            slot_usado = ""

            if override and "responsables" in override:
                responsable_control = normalizar_lista_profes(override.get("responsables", []))
                slot_usado = "MANUAL"

            elif override and "slot" in override:
                slot_override = str(override.get("slot", "")).strip().upper()
                encontrado = False

                for item in pool_slots:
                    if str(item.get("slot", "")).strip().upper() == slot_override:
                        responsable_control = [str(item.get("profesor", "")).strip()]
                        slot_usado = slot_override
                        encontrado = True
                        break

                if not encontrado:
                    elegido = elegir_balanceado(pool_control, carga_actual, horas_map, cantidad=1)
                    if elegido:
                        responsable_control = elegido
                        slot_usado = "BALANCEADO"

            else:
                elegido = elegir_balanceado(pool_control, carga_actual, horas_map, cantidad=1)
                if elegido:
                    responsable_control = elegido
                    slot_usado = "BALANCEADO"

            filas_chequeo.append({
                "sección": seccion,
                "fecha_evento": fecha_evento.date(),
                "evento": nombre,
                "observaciones": obs_cal,
                "slot_pool": slot_usado,
                "responsable_control": normalizar_profes_str(responsable_control),
            })

            for paso_key, paso in proto.items():
                offset = int(paso.get("offset_dias", 0))
                deadline = (fecha_evento + pd.Timedelta(days=offset)).date()

                detalle = str(paso.get("detalle", "")).strip()
                if detalle_base_evento:
                    detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

                filas.append({
                    "fecha_limite": deadline,
                    "fecha_evento": fecha_evento.date(),
                    "evento": nombre,
                    "tipo_evento": proto_nombre,
                    "paso": paso_key,
                    "sección": seccion,
                    "responsables": normalizar_profes_str(responsable_control),
                    "detalle": detalle,
                    "estado": "Pendiente",
                })

                # contar esta carga como repartible de seminario
                if paso_key in ["construir_control", "escanear", "corregir_y_notas"]:
                    sumar_carga(carga_actual, responsable_control, peso=1.0)

            continue


        if proto_nombre == "Portafolio":
            for paso_key, paso in proto.items():
                offset = int(paso.get("offset_dias", 0))
                deadline = (fecha_evento + pd.Timedelta(days=offset)).date()

                detalle = str(paso.get("detalle", "")).strip()
                if detalle_base_evento:
                    detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

                filas.append({
                    "fecha_limite": deadline,
                    "fecha_evento": fecha_evento.date(),
                    "evento": nombre,
                    "tipo_evento": proto_nombre,
                    "paso": paso_key,
                    "sección": seccion,
                    "responsables": normalizar_profes_str(participantes),
                    "detalle": detalle,
                    "estado": "Pendiente",
                })
            continue


        # ========================================================
        # TALLER
        # pool global ponderado, pero sin repetir persona dentro del mismo taller
        # ========================================================
        # if proto_nombre == "Taller":
        #     pool_global = obtener_pool_global_participantes(config)

        #     if len(pool_global) == 0:
        #         continue

        #     elegidos = elegir_varios_ciclico_sin_repetir(pool_global, contador_taller_global, 4)

        #     if len(elegidos) < 4:
        #         faltan = 4 - len(elegidos)
        #         extra = elegir_varios_ciclico_sin_repetir(pool_global, contador_taller_global + 4, 4 + faltan)
        #         for p in extra:
        #             if p not in elegidos:
        #                 elegidos.append(p)
        #             if len(elegidos) == 4:
        #                 break

        #     if len(elegidos) < 4:
        #         continue

        #     contador_taller_global += 4

        #     p1, p2, p3, p4 = elegidos[:4]

        #     for paso_key, paso in proto.items():
        #         offset = int(paso.get("offset_dias", 0))
        #         deadline = (fecha_evento + pd.Timedelta(days=offset)).date()
        #         detalle = str(paso.get("detalle", "")).strip()
        #         if detalle_base_evento:
        #             detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

        #         if paso_key == "construir_taller":
        #             filas.append({
        #                 "fecha_limite": deadline,
        #                 "fecha_evento": fecha_evento.date(),
        #                 "evento": nombre,
        #                 "tipo_evento": proto_nombre,
        #                 "paso": "construir_taller_AB",
        #                 "sección": seccion,
        #                 "responsables": normalizar_profes_str([p1, p2]),
        #                 "detalle": f"{detalle} — Versiones A y B.",
        #                 "estado": "Pendiente",
        #             })
        #             filas.append({
        #                 "fecha_limite": deadline,
        #                 "fecha_evento": fecha_evento.date(),
        #                 "evento": nombre,
        #                 "tipo_evento": proto_nombre,
        #                 "paso": "construir_taller_CD",
        #                 "sección": seccion,
        #                 "responsables": normalizar_profes_str([p3, p4]),
        #                 "detalle": f"{detalle} — Versiones C y D.",
        #                 "estado": "Pendiente",
        #             })

        #         elif paso_key == "corregir_taller":
        #             for prof, version in [(p1, "A"), (p2, "B"), (p3, "C"), (p4, "D")]:
        #                 filas.append({
        #                     "fecha_limite": deadline,
        #                     "fecha_evento": fecha_evento.date(),
        #                     "evento": nombre,
        #                     "tipo_evento": proto_nombre,
        #                     "paso": f"corregir_taller_{version}",
        #                     "sección": seccion,
        #                     "responsables": normalizar_profes_str([prof]),
        #                     "detalle": f"{detalle} — Corregir versión {version}.",
        #                     "estado": "Pendiente",
        #                 })

        #         else:
        #             filas.append({
        #                 "fecha_limite": deadline,
        #                 "fecha_evento": fecha_evento.date(),
        #                 "evento": nombre,
        #                 "tipo_evento": proto_nombre,
        #                 "paso": paso_key,
        #                 "sección": seccion,
        #                 "responsables": normalizar_profes_str(elegidos),
        #                 "detalle": detalle,
        #                 "estado": "Pendiente",
        #             })
        #     continue


        # if proto_nombre == "Taller":
        #     pool_global = obtener_pool_global_participantes(config)

        #     if len(pool_global) == 0:
        #         continue

        #     elegidos, nuevo_cursor = elegir_varios_ciclico_sin_repetir(
        #         pool_global,
        #         contador_taller_global,
        #         4
        #     )

        #     if len(elegidos) < 4:
        #         elegidos_extra, cursor_extra = elegir_varios_ciclico_sin_repetir(
        #             pool_global,
        #             nuevo_cursor,
        #             4
        #         )
        #         for p in elegidos_extra:
        #             if p not in elegidos:
        #                 elegidos.append(p)
        #             if len(elegidos) == 4:
        #                 break
        #         nuevo_cursor = cursor_extra

        #     if len(elegidos) < 4:
        #         continue

        #     contador_taller_global = nuevo_cursor

        #     p1, p2, p3, p4 = elegidos[:4]
        
        
        
        
        
        
        
        # if proto_nombre == "Taller":
        #     pool_global = obtener_pool_global_participantes(config)

        #     if len(pool_global) == 0:
        #         continue

        #     elegidos, nuevo_cursor = elegir_varios_ciclico_sin_repetir(
        #         pool_global,
        #         contador_taller_global,
        #         4
        #     )

        #     if len(elegidos) < 4:
        #         elegidos_extra, cursor_extra = elegir_varios_ciclico_sin_repetir(
        #             pool_global,
        #             nuevo_cursor,
        #             4
        #         )
        #         for p in elegidos_extra:
        #             if p not in elegidos:
        #                 elegidos.append(p)
        #             if len(elegidos) == 4:
        #                 break
        #         nuevo_cursor = cursor_extra

        #     if len(elegidos) < 4:
        #         continue

        #     contador_taller_global = nuevo_cursor

        #     p1, p2, p3, p4 = elegidos[:4]

        #     for paso_key, paso in proto.items():
        #         offset = int(paso.get("offset_dias", 0))
        #         deadline = (fecha_evento + pd.Timedelta(days=offset)).date()

        #         detalle = str(paso.get("detalle", "")).strip()
        #         if detalle_base_evento:
        #             detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

        #         if paso_key == "construir_taller":
        #             filas.append({
        #                 "fecha_limite": deadline,
        #                 "fecha_evento": fecha_evento.date(),
        #                 "evento": nombre,
        #                 "tipo_evento": proto_nombre,
        #                 "paso": "construir_taller_AB",
        #                 "sección": seccion,
        #                 "responsables": normalizar_profes_str([p1, p2]),
        #                 "detalle": f"{detalle} — Versiones A y B.",
        #                 "estado": "Pendiente",
        #             })

        #             filas.append({
        #                 "fecha_limite": deadline,
        #                 "fecha_evento": fecha_evento.date(),
        #                 "evento": nombre,
        #                 "tipo_evento": proto_nombre,
        #                 "paso": "construir_taller_CD",
        #                 "sección": seccion,
        #                 "responsables": normalizar_profes_str([p3, p4]),
        #                 "detalle": f"{detalle} — Versiones C y D.",
        #                 "estado": "Pendiente",
        #             })

        #         elif paso_key == "corregir_taller":
        #             for prof, version in [(p1, "A"), (p2, "B"), (p3, "C"), (p4, "D")]:
        #                 filas.append({
        #                     "fecha_limite": deadline,
        #                     "fecha_evento": fecha_evento.date(),
        #                     "evento": nombre,
        #                     "tipo_evento": proto_nombre,
        #                     "paso": f"corregir_taller_{version}",
        #                     "sección": seccion,
        #                     "responsables": normalizar_profes_str([prof]),
        #                     "detalle": f"{detalle} — Corregir versión {version}.",
        #                     "estado": "Pendiente",
        #                 })

        #         else:
        #             filas.append({
        #                 "fecha_limite": deadline,
        #                 "fecha_evento": fecha_evento.date(),
        #                 "evento": nombre,
        #                 "tipo_evento": proto_nombre,
        #                 "paso": paso_key,
        #                 "sección": seccion,
        #                 "responsables": normalizar_profes_str(elegidos),
        #                 "detalle": detalle,
        #                 "estado": "Pendiente",
        #             })

        #     continue
        
        
        
        # if proto_nombre == "Taller":
        #     pool_global = obtener_pool_global_participantes(config)

        #     if len(pool_global) == 0:
        #         continue

        #     # una persona para construir
        #     elegido_construccion, cursor_1 = elegir_varios_ciclico_sin_repetir(
        #         pool_global,
        #         contador_taller_global,
        #         1
        #     )

        #     if len(elegido_construccion) < 1:
        #         continue

        #     # una persona para corregir (idealmente distinta)
        #     elegido_correccion, cursor_2 = elegir_varios_ciclico_sin_repetir(
        #         pool_global,
        #         cursor_1,
        #         1
        #     )

        #     if len(elegido_correccion) < 1:
        #         elegido_correccion = elegido_construccion[:]
        #         cursor_2 = cursor_1

        #     profe_construccion = elegido_construccion[0]
        #     profe_correccion = elegido_correccion[0]

        #     # overrides manuales
        #     override_construccion = buscar_override(
        #         config,
        #         proto_nombre,
        #         seccion,
        #         fecha_evento.date(),
        #         paso="construir_taller"
        #     )
        #     if override_construccion and "responsables" in override_construccion:
        #         manual = normalizar_lista_profes(override_construccion.get("responsables", []))
        #         if manual:
        #             profe_construccion = manual[0]

        #     override_correccion = buscar_override(
        #         config,
        #         proto_nombre,
        #         seccion,
        #         fecha_evento.date(),
        #         paso="corregir_taller"
        #     )
        #     if override_correccion and "responsables" in override_correccion:
        #         manual = normalizar_lista_profes(override_correccion.get("responsables", []))
        #         if manual:
        #             profe_correccion = manual[0]

        #     contador_taller_global = cursor_2

        #     for paso_key, paso in proto.items():
        #         offset = int(paso.get("offset_dias", 0))
        #         deadline = (fecha_evento + pd.Timedelta(days=offset)).date()

        #         detalle = str(paso.get("detalle", "")).strip()
        #         if detalle_base_evento:
        #             detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

        #         if paso_key == "construir_taller":
        #             responsables = [profe_construccion]
        #             detalle_final = detalle

        #         elif paso_key == "corregir_taller":
        #             responsables = [profe_correccion]
        #             detalle_final = detalle

        #         else:
        #             responsables = [profe_construccion]
        #             detalle_final = detalle

        #         filas.append({
        #             "fecha_limite": deadline,
        #             "fecha_evento": fecha_evento.date(),
        #             "evento": nombre,
        #             "tipo_evento": proto_nombre,
        #             "paso": paso_key,
        #             "sección": seccion,
        #             "responsables": normalizar_profes_str(responsables),
        #             "detalle": detalle_final,
        #             "estado": "Pendiente",
        #         })

        #     continue



        if proto_nombre == "Taller":
            pool_global = obtener_pool_global_participantes(config)

            if len(pool_global) == 0:
                continue

            elegido_construccion = elegir_balanceado(
                pool_global,
                carga_actual,
                horas_map,
                cantidad=1
            )

            if len(elegido_construccion) < 1:
                continue

            elegido_correccion = elegir_balanceado(
                pool_global,
                carga_actual,
                horas_map,
                cantidad=1,
                excluir=set(elegido_construccion)
            )

            if len(elegido_correccion) < 1:
                elegido_correccion = elegido_construccion[:]

            profe_construccion = elegido_construccion[0]
            profe_correccion = elegido_correccion[0]

            override_construccion = buscar_override(
                config,
                proto_nombre,
                seccion,
                fecha_evento.date(),
                paso="construir_taller"
            )
            if override_construccion and "responsables" in override_construccion:
                manual = normalizar_lista_profes(override_construccion.get("responsables", []))
                if manual:
                    profe_construccion = manual[0]

            override_correccion = buscar_override(
                config,
                proto_nombre,
                seccion,
                fecha_evento.date(),
                paso="corregir_taller"
            )
            if override_correccion and "responsables" in override_correccion:
                manual = normalizar_lista_profes(override_correccion.get("responsables", []))
                if manual:
                    profe_correccion = manual[0]

            for paso_key, paso in proto.items():
                offset = int(paso.get("offset_dias", 0))
                deadline = (fecha_evento + pd.Timedelta(days=offset)).date()

                detalle = str(paso.get("detalle", "")).strip()
                if detalle_base_evento:
                    detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

                if paso_key == "construir_taller":
                    responsables = [profe_construccion]
                elif paso_key == "corregir_taller":
                    responsables = [profe_correccion]
                else:
                    responsables = [profe_construccion]

                filas.append({
                    "fecha_limite": deadline,
                    "fecha_evento": fecha_evento.date(),
                    "evento": nombre,
                    "tipo_evento": proto_nombre,
                    "paso": paso_key,
                    "sección": seccion,
                    "responsables": normalizar_profes_str(responsables),
                    "detalle": detalle,
                    "estado": "Pendiente",
                })

                if paso_key in ["construir_taller", "corregir_taller"]:
                    sumar_carga(carga_actual, responsables, peso=1.0)

            continue



        # ========================================================
        # CERTAMEN
        # pedir_preguntas: participantes
        # construir/pauta: PEC
        # corregir_y_notas: participantes
        # ========================================================
        if proto_nombre == "Certamen":
            for paso_key, paso in proto.items():
                offset = int(paso.get("offset_dias", 0))
                deadline = (fecha_evento + pd.Timedelta(days=offset)).date()
                detalle = str(paso.get("detalle", "")).strip()
                if detalle_base_evento:
                    detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

                if paso_key == "pedir_preguntas":
                    responsables = participantes
                elif paso_key in ["construir_control", "pauta_prueba"]:
                    responsables = [pec] if pec else []
                elif paso_key == "corregir_y_notas":
                    responsables = participantes
                else:
                    responsables = participantes

                filas.append({
                    "fecha_limite": deadline,
                    "fecha_evento": fecha_evento.date(),
                    "evento": nombre,
                    "tipo_evento": proto_nombre,
                    "paso": paso_key,
                    "sección": seccion,
                    "responsables": normalizar_profes_str(responsables),
                    "detalle": detalle,
                    "estado": "Pendiente",
                })
            continue

        # ========================================================
        # EXAMEN
        # pedir_preguntas: participantes
        # construir/pauta: PEC
        # corregir_examen: participantes
        # ========================================================
        if proto_nombre == "Examen":
            for paso_key, paso in proto.items():
                offset = int(paso.get("offset_dias", 0))
                deadline = (fecha_evento + pd.Timedelta(days=offset)).date()
                detalle = str(paso.get("detalle", "")).strip()
                if detalle_base_evento:
                    detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

                if paso_key == "pedir_preguntas":
                    responsables = participantes
                elif paso_key in ["construir_examen", "pauta_examen"]:
                    responsables = [pec] if pec else []
                elif paso_key == "corregir_examen":
                    responsables = participantes
                else:
                    responsables = participantes

                filas.append({
                    "fecha_limite": deadline,
                    "fecha_evento": fecha_evento.date(),
                    "evento": nombre,
                    "tipo_evento": proto_nombre,
                    "paso": paso_key,
                    "sección": seccion,
                    "responsables": normalizar_profes_str(responsables),
                    "detalle": detalle,
                    "estado": "Pendiente",
                })
            continue

        # ========================================================
        # TRABAJO PRÁCTICO
        # ========================================================
        if proto_nombre == "Trabajo práctico":
            for paso_key, paso in proto.items():
                offset = int(paso.get("offset_dias", 0))
                deadline = (fecha_evento + pd.Timedelta(days=offset)).date()
                detalle = str(paso.get("detalle", "")).strip()
                if detalle_base_evento:
                    detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

                filas.append({
                    "fecha_limite": deadline,
                    "fecha_evento": fecha_evento.date(),
                    "evento": nombre,
                    "tipo_evento": proto_nombre,
                    "paso": paso_key,
                    "sección": seccion,
                    "responsables": normalizar_profes_str(participantes),
                    "detalle": detalle,
                    "estado": "Pendiente",
                })
            continue

        # ========================================================
        # LABORATORIO
        # ========================================================
        if proto_nombre == "Laboratorio":
            idx_lab = contador_lab_video_por_seccion.get(seccion, 0)
            video_lab = elegir_ciclico(pool_lab, idx_lab)
            contador_lab_video_por_seccion[seccion] = idx_lab + 1

            for paso_key, paso in proto.items():
                offset = int(paso.get("offset_dias", 0))
                deadline = (fecha_evento + pd.Timedelta(days=offset)).date()
                detalle = str(paso.get("detalle", "")).strip()
                if detalle_base_evento:
                    detalle = f"{detalle} — {detalle_base_evento}" if detalle else detalle_base_evento

                if paso_key == "preparar_material_previo":
                    responsables = [pec] if pec else []
                elif paso_key == "grabar_video_solucion":
                    responsables = video_lab
                elif paso_key == "corregir_informe_laboratorio":
                    responsables = pool_lab
                else:
                    responsables = pool_lab

                filas.append({
                    "fecha_limite": deadline,
                    "fecha_evento": fecha_evento.date(),
                    "evento": nombre,
                    "tipo_evento": proto_nombre,
                    "paso": paso_key,
                    "sección": seccion,
                    "responsables": normalizar_profes_str(responsables),
                    "detalle": detalle,
                    "estado": "Pendiente",
                })
            continue

    df_mis = pd.DataFrame(filas)

    if df_mis.empty:
        columnas = [
            "fecha_limite", "fecha_evento", "evento", "tipo_evento", "paso",
            "sección", "responsables", "detalle", "estado"
        ]
        return (
            pd.DataFrame(columns=columnas),
            pd.DataFrame(filas_chequeo),
            pd.DataFrame(filas_pools)
        )

    df_mis = df_mis.sort_values(
        ["fecha_limite", "fecha_evento", "sección", "tipo_evento", "paso"]
    ).reset_index(drop=True)

    df_chequeo = pd.DataFrame(filas_chequeo)
    df_pools = pd.DataFrame(filas_pools)

    return df_mis, df_chequeo, df_pools


# ============================================================
# MATRIZ
# ============================================================
def armar_matriz(df_mis, config):
    alias = config.get("misiones", {}).get("alias_seccion", {}) or {}

    if df_mis.empty:
        return pd.DataFrame()

    df = df_mis.copy()

    for c in ["evento", "tipo_evento", "paso", "detalle", "sección", "responsables"]:
        if c not in df.columns:
            df[c] = ""
        df[c] = df[c].fillna("").astype(str)

    df["sección_col"] = df["sección"].apply(lambda s: alias.get(s, s))

    mat = df.pivot_table(
        index=["evento", "tipo_evento", "paso", "detalle"],
        columns="sección_col",
        values="responsables",
        aggfunc="first",
        fill_value=""
    ).reset_index()

    mat = mat.rename(columns={
        "evento": "Evaluación",
        "tipo_evento": "Tipo",
        "paso": "Misión",
        "detalle": "Detalle",
    })

    return mat


# ============================================================
# EXPORTAR
# ============================================================
def style_sheet(ws, header_color="1F4E78"):
    fill_header = PatternFill("solid", fgColor=header_color)
    font_header = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for col in range(1, ws.max_column + 1):
        maxlen = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, maxlen + 2), 60)


def exportar_excel(df_mis, df_mat, df_chequeo, df_pools, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
        df_mis.to_excel(writer, index=False, sheet_name="Misiones")

        if not df_mat.empty:
            df_mat.to_excel(writer, index=False, sheet_name="Matriz")

        if not df_chequeo.empty:
            df_chequeo.to_excel(writer, index=False, sheet_name="Chequeo")

        if not df_pools.empty:
            df_pools.to_excel(writer, index=False, sheet_name="Pools")

        orden = [
            "pedir_preguntas",
            "construir_control",
            "pauta_prueba",
            "revisar_prueba",
            "escanear",
            "corregir_y_notas",
            "revisar_tp",
            # "construir_taller_AB",
            # "construir_taller_CD",
            # "corregir_taller_A",
            # "corregir_taller_B",
            # "corregir_taller_C",
            # "corregir_taller_D",
            "construir_taller",
            "corregir_taller",
            "construir_examen",
            "pauta_examen",
            "corregir_examen",
            "preparar_material_previo",
            "grabar_video_solucion",
            "corregir_informe_laboratorio",
            "revisar_portafolio",
        ]

        df_plan = df_mis.copy()
        df_plan["paso_rank"] = df_plan["paso"].apply(lambda x: orden.index(x) if x in orden else 999)
        df_plan = df_plan.sort_values(
            ["fecha_evento", "sección", "tipo_evento", "paso_rank", "fecha_limite"]
        ).drop(columns=["paso_rank"])
        df_plan.to_excel(writer, index=False, sheet_name="Plan")

        wb = writer.book

        style_sheet(wb["Misiones"], "5B2C6F")
        style_sheet(wb["Plan"], "2C3E50")

        if "Matriz" in wb.sheetnames:
            style_sheet(wb["Matriz"], "7F7F7F")

        if "Chequeo" in wb.sheetnames:
            style_sheet(wb["Chequeo"], "0F766E")

        if "Pools" in wb.sheetnames:
            style_sheet(wb["Pools"], "92400E")

        ws_mis = wb["Misiones"]
        headers = [c.value for c in ws_mis[1]]

        if "fecha_limite" in headers:
            idx = headers.index("fecha_limite") + 1
            fill_deadline = PatternFill("solid", fgColor="F4CCCC")
            for r in range(2, ws_mis.max_row + 1):
                ws_mis.cell(r, idx).fill = fill_deadline
                ws_mis.cell(r, idx).alignment = Alignment(horizontal="center")


# ============================================================
# MAIN
# ============================================================
def main():
    cursos = {
        "fokito": {
            "config_path": os.path.join("config", "calendario_fokito.yml"),
            "cal_path": os.path.join("data", "fokito", "calendario.xlsx"),
            "out_path": os.path.join("data", "fokito", "misiones.xlsx"),
        },
        "tecnologia_medica": {
            "config_path": os.path.join("config", "calendario_tecnologia_medica.yml"),
            "cal_path": os.path.join("data", "tecnologia_medica", "calendario.xlsx"),
            "out_path": os.path.join("data", "tecnologia_medica", "misiones.xlsx"),
        },
        "medicina": {
            "config_path": os.path.join("config", "calendario_medicina.yml"),
            "cal_path": os.path.join("data", "medicina", "calendario.xlsx"),
            "out_path": os.path.join("data", "medicina", "misiones.xlsx"),
        },
        "enobnu": {
            "config_path": os.path.join("config", "calendario_enobnu.yml"),
            "cal_path": os.path.join("data", "enobnu", "calendario.xlsx"),
            "out_path": os.path.join("data", "enobnu", "misiones.xlsx"),
        },
    }

    for curso, info in cursos.items():
        config_path = info["config_path"]
        cal_path = info["cal_path"]
        out_path = info["out_path"]

        if not os.path.exists(config_path):
            print(f"⚠️  Saltando {curso}: no existe {config_path}")
            continue

        if not os.path.exists(cal_path):
            print(f"⚠️  Saltando {curso}: no existe {cal_path}. Genera primero el calendario.")
            continue

        config = cargar_yaml(config_path)
        df_cal = pd.read_excel(cal_path, sheet_name="Calendario")

        estados_previos = cargar_estados_previos(out_path)

        df_mis, df_chequeo, df_pools = construir_misiones(config, df_cal)
        df_mis = aplicar_estados_previos(df_mis, estados_previos)
        df_mat = armar_matriz(df_mis, config)


        graficar_equilibrio_curso(
            nombre_curso=curso,
            df_cal=df_cal,
            df_mis=df_mis,
            carpeta_salida=os.path.dirname(out_path)
        )


        exportar_excel(df_mis, df_mat, df_chequeo, df_pools, out_path)

        print(f"OK: [{curso}] generado {out_path} con {len(df_mis)} tareas.")

        if not df_chequeo.empty:
            print("  Controles asignados:")
            print(
                df_chequeo[
                    ["sección", "fecha_evento", "slot_pool", "responsable_control"]
                ].to_string(index=False)
            )


if __name__ == "__main__":
    main()